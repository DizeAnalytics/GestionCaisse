import os
from io import BytesIO
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from django.conf import settings
from django.http import HttpResponse
from datetime import datetime
from reportlab.pdfgen import canvas
import logging
from django.utils import timezone
from .models import Parametre
from reportlab.platypus import Table as RLTable, TableStyle as RLTableStyle
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.barcharts import VerticalBarChart
from django.core.files.base import ContentFile
from reportlab.platypus import PageBreak

logger = logging.getLogger(__name__)


def validate_image_file(image_field):
    """
    Valide qu'un champ d'image est valide et accessible
    """
    if not image_field:
        return False
    
    try:
        # Vérifier que le champ a un attribut path
        if not hasattr(image_field, 'path'):
            return False
        
        # Vérifier que le fichier existe
        if not os.path.exists(image_field.path):
            return False
        
        # Vérifier que c'est un fichier (pas un dossier)
        if not os.path.isfile(image_field.path):
            return False
        
        # Vérifier que le fichier n'est pas vide
        if os.path.getsize(image_field.path) == 0:
            return False
        
        return True
    except Exception as e:
        logger.warning(f"Erreur lors de la validation de l'image: {e}")
        return False


def get_parametres_application():
    """
    Récupère les paramètres actifs de l'application.
    Retourne un dictionnaire avec les valeurs par défaut si aucun paramètre n'est configuré.
    """
    try:
        parametres = Parametre.get_parametres_actifs()
        if parametres:
            return {
                'nom_application': parametres.nom_application,
                'logo': parametres.logo,
                'description_application': parametres.description_application,
                'version_application': parametres.version_application,
                'telephone_principal': parametres.telephone_principal,
                'telephone_secondaire': parametres.telephone_secondaire,
                'email_contact': parametres.email_contact,
                'site_web': parametres.site_web,
                'siege_social': parametres.siege_social,
                'adresse_postale': parametres.adresse_postale,
                'boite_postale': parametres.boite_postale,
                'ville': parametres.ville,
                'pays': parametres.pays,
                'nom_president_general': parametres.nom_president_general,
                'titre_president_general': parametres.titre_president_general,
                'signature_president_general': parametres.signature_president_general,
                'nom_directeur_technique': parametres.nom_directeur_technique,
                'nom_directeur_financier': parametres.nom_directeur_financier,
                'nom_directeur_administratif': parametres.nom_directeur_administratif,
                'numero_agrement': parametres.numero_agrement,
                'date_agrement': parametres.date_agrement,
                'autorite_agrement': parametres.autorite_agrement,
                'devise': parametres.devise,
                'langue_par_defaut': parametres.langue_par_defaut,
                'fuseau_horaire': parametres.fuseau_horaire,
                'copyright_text': parametres.copyright_text,
                'mentions_legales': parametres.mentions_legales,
            }
    except Exception as e:
        logger.warning(f"Erreur lors de la récupération des paramètres: {e}")
    
    # Valeurs par défaut si aucun paramètre n'est configuré
    return {
        'nom_application': 'CAISSE DE SOLIDARITÉ',
        'logo': None,
        'description_application': '',
        'version_application': '1.0.0',
        'telephone_principal': '',
        'telephone_secondaire': '',
        'email_contact': '',
        'site_web': '',
        'siege_social': '',
        'adresse_postale': '',
        'boite_postale': '',
        'ville': '',
        'pays': 'Togo',
        'nom_president_general': '',
        'titre_president_general': 'Président Général',
        'signature_president_general': None,
        'nom_directeur_technique': '',
        'nom_directeur_financier': '',
        'nom_directeur_administratif': '',
        'numero_agrement': '',
        'date_agrement': None,
        'autorite_agrement': '',
        'devise': 'FCFA',
        'langue_par_defaut': 'fr',
        'fuseau_horaire': 'Africa/Lome',
        'copyright_text': '',
        'mentions_legales': '',
    }


def get_signature_president_general():
    """
    Récupère les informations de signature du Président Général depuis les paramètres.
    Retourne un tuple (nom, titre, signature_image, signature_available)
    """
    parametres = get_parametres_application()
    
    nom = parametres.get('nom_president_general', 'Non défini')
    titre = parametres.get('titre_president_general', 'Président Général')
    signature = parametres.get('signature_president_general')
    
    signature_available = False
    signature_image = None
    
    if nom and nom != 'Non défini' and signature and validate_image_file(signature):
        try:
            signature_image = Image(signature.path, width=1*inch, height=0.5*inch)
            signature_available = True
        except Exception as e:
            logger.warning(f"Erreur lors du chargement de la signature du président général: {e}")
            signature_image = ""
    else:
        signature_image = ""
    
    return nom, titre, signature_image, signature_available


def create_signatures_table_with_demandeur_first(pret_or_membre, signatures_data):
    """
    Crée un tableau de signatures avec l'ordre spécifié :
    1. Demandeur (Membre bénéficiaire)
    2. Trésorière de la caisse
    3. Secrétaire de la caisse
    4. Présidente de la caisse
    5. PCA de toutes les caisses
    """
    # 1. SIGNATURE DU DEMANDEUR (Membre) - EN PREMIER
    if hasattr(pret_or_membre, 'membre') and pret_or_membre.membre:
        # Cas d'un prêt
        membre = pret_or_membre.membre
    elif hasattr(pret_or_membre, 'nom_complet'):
        # Cas d'un membre direct
        membre = pret_or_membre
    else:
        membre = None
    
    # Créer un nouveau tableau avec l'ordre correct
    new_signatures_data = []
    
    # 1. Demandeur (Membre bénéficiaire) - EN PREMIER
    if membre:
        if validate_image_file(membre.signature):
            try:
                sig_membre = Image(membre.signature.path, width=1*inch, height=0.5*inch)
            except Exception as e:
                logger.warning(f"Erreur lors du chargement de la signature du membre: {e}")
                sig_membre = ""
        else:
            sig_membre = ""
        
        new_signatures_data.append([
            "Demandeur (Membre bénéficiaire):",
            sig_membre,
            membre.nom_complet
        ])
    else:
        new_signatures_data.append([
            "Demandeur (Membre bénéficiaire):",
            "Non défini",
            "Non défini"
        ])
    
    # 2. Trésorière de la caisse
    tresoriere_found = False
    for sig_data in signatures_data:
        if "Trésorière" in sig_data[0]:
            new_signatures_data.append(sig_data)
            tresoriere_found = True
            break
    
    if not tresoriere_found:
        new_signatures_data.append([
            "Trésorière de la caisse:",
            "Non définie",
            "Non définie"
        ])
    
    # 3. Secrétaire de la caisse
    secretaire_found = False
    for sig_data in signatures_data:
        if "Secrétaire" in sig_data[0]:
            new_signatures_data.append(sig_data)
            secretaire_found = True
            break
    
    if not secretaire_found:
        new_signatures_data.append([
            "Secrétaire de la caisse:",
            "Non définie",
            "Non définie"
        ])
    
    # 4. Présidente de la caisse
    presidente_found = False
    for sig_data in signatures_data:
        if "Présidente" in sig_data[0]:
            new_signatures_data.append(sig_data)
            presidente_found = True
            break
    
    if not presidente_found:
        new_signatures_data.append([
            "Présidente de la caisse:",
            "Non définie",
            "Non définie"
        ])
    
    # 5. PCA de toutes les caisses - EN DERNIER
    pca_found = False
    for sig_data in signatures_data:
        if "PCA" in sig_data[0] or "Président Général" in sig_data[0]:
            new_signatures_data.append(sig_data)
            pca_found = True
            break
    
    if not pca_found:
        new_signatures_data.append([
            "PCA de toutes les caisses:",
            "Non défini",
            "Non défini"
        ])
    
    return new_signatures_data


def add_contact_info_to_pdf(story, parametres):
    """
    Ajoute les informations de contact depuis les paramètres au PDF.
    """
    if not parametres:
        parametres = get_parametres_application()
    
    # Styles
    styles = getSampleStyleSheet()
    
    # Style de section
    section_style = ParagraphStyle(
        'Section',
        parent=styles['Heading3'],
        fontSize=14,
        spaceAfter=15,
        spaceBefore=20,
        textColor=colors.HexColor('#F18F01'),
        fontName='Helvetica-Bold'
    )
    
    # Créer une section pour les informations de contact
    contact_info = []
    
    # Téléphones
    if parametres.get('telephone_principal'):
        contact_info.append(f"📞 Téléphone: {parametres['telephone_principal']}")
    if parametres.get('telephone_secondaire'):
        contact_info.append(f"📞 Téléphone secondaire: {parametres['telephone_secondaire']}")
    
    # Email
    if parametres.get('email_contact'):
        contact_info.append(f"📧 Email: {parametres['email_contact']}")
    
    # Site web
    if parametres.get('site_web'):
        contact_info.append(f"🌐 Site web: {parametres['site_web']}")
    
    # Adresse
    if parametres.get('siege_social'):
        contact_info.append(f"📍 Siège social: {parametres['siege_social']}")
    if parametres.get('adresse_postale'):
        contact_info.append(f"📮 Adresse postale: {parametres['adresse_postale']}")
    if parametres.get('boite_postale'):
        contact_info.append(f"📮 Boîte postale: {parametres['boite_postale']}")
    if parametres.get('ville'):
        contact_info.append(f"🏙️ Ville: {parametres['ville']}")
    if parametres.get('pays'):
        contact_info.append(f"🌍 Pays: {parametres['pays']}")
    
    # Ajouter les informations de contact si disponibles
    if contact_info:
        story.append(Spacer(1, 20))
        story.append(Paragraph("📞 INFORMATIONS DE CONTACT", section_style))
        
        contact_style = ParagraphStyle(
            'ContactInfo',
            parent=styles['Normal'],
            fontSize=9,
            spaceAfter=5,
            alignment=TA_LEFT,
            textColor=colors.HexColor('#666666'),
            fontName='Helvetica'
        )
        
        for info in contact_info:
            story.append(Paragraph(info, contact_style))


def create_standard_header(story, parametres, title=None, subtitle=None):
    """
    Crée un en-tête standard pour tous les PDFs avec logo, nom de l'application et informations du PDG
    """
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Title'],
        fontSize=18,
        spaceAfter=12,
        alignment=TA_CENTER,
        textColor=colors.darkblue
    )
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Heading2'],
        fontSize=14,
        spaceAfter=20,
        alignment=TA_CENTER,
        textColor=colors.darkblue
    )
    header_style = ParagraphStyle(
        'Header',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=6,
        alignment=TA_CENTER,
        textColor=colors.grey
    )
    
    # En-tête avec logo et nom de l'application
    if validate_image_file(parametres['logo']):
        try:
            logo_img = Image(parametres['logo'].path, width=1.5*inch, height=1.5*inch)
            story.append(logo_img)
            story.append(Spacer(1, 10))
        except Exception as e:
            logger.warning(f"Erreur lors du chargement du logo: {e}")
            # Continuer sans le logo
    
    # Nom de l'application
    story.append(Paragraph(f"<b>{parametres['nom_application']}</b>", title_style))
    
    # Description de l'application si disponible
    if parametres['description_application']:
        story.append(Paragraph(parametres['description_application'], header_style))
    
    # Informations de contact
    contact_info = []
    if parametres['telephone_principal']:
        contact_info.append(f"Tél: {parametres['telephone_principal']}")
    if parametres['email_contact']:
        contact_info.append(f"Email: {parametres['email_contact']}")
    if parametres['siege_social']:
        contact_info.append(f"Siège: {parametres['siege_social']}")
    
    if contact_info:
        story.append(Paragraph(" | ".join(contact_info), header_style))
    
    story.append(Spacer(1, 20))
    
    # Titre du document
    if title:
        story.append(Paragraph(title, subtitle_style))
    if subtitle:
        story.append(Paragraph(subtitle, header_style))
    
    story.append(Spacer(1, 20))


def create_standard_footer(story, parametres):
    """
    Crée un pied de page standard pour tous les PDFs avec les informations du PDG
    """
    styles = getSampleStyleSheet()
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=9,
        spaceAfter=6,
        alignment=TA_CENTER,
        textColor=colors.grey
    )
    
    # Informations du PDG
    if parametres['nom_president_general']:
        story.append(Spacer(1, 30))
        story.append(Paragraph(f"<b>{parametres['titre_president_general']}:</b> {parametres['nom_president_general']}", footer_style))
    
    # Copyright et mentions légales
    if parametres['copyright_text']:
        story.append(Paragraph(parametres['copyright_text'], footer_style))
    
    # Date de génération
    story.append(Paragraph(f"Document généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}", footer_style))


def generate_credentials_pdf(caisse, created_users):
    """
    Génère un PDF avec les informations des comptes utilisateurs créés
    """
    try:
        # Créer le buffer pour le PDF
        buffer = BytesIO()
        
        # Créer le document PDF
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        story = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=TA_CENTER,
            textColor=colors.darkblue
        )
        
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Heading2'],
            fontSize=14,
            spaceAfter=20,
            alignment=TA_CENTER,
            textColor=colors.darkgreen
        )
        
        normal_style = styles['Normal']
        bold_style = ParagraphStyle(
            'Bold',
            parent=styles['Normal'],
            fontSize=12,
            spaceAfter=6,
            textColor=colors.black
        )
        
        # Titre principal
        story.append(Paragraph("INFORMATIONS DE CONNEXION DES RESPONSABLES", title_style))
        story.append(Spacer(1, 20))
        
        # Informations de la caisse
        story.append(Paragraph("INFORMATIONS DE LA CAISSE", subtitle_style))
        
        caisse_nom = getattr(caisse, 'nom_association', '') or ''
        caisse_code = getattr(caisse, 'code', '') or ''
        caisse_date_creation = getattr(caisse, 'date_creation', None)
        try:
            caisse_date_str = caisse_date_creation.strftime('%d/%m/%Y à %H:%M') if caisse_date_creation else ''
        except Exception:
            caisse_date_str = ''
        try:
            statut_str = caisse.get_statut_display() if hasattr(caisse, 'get_statut_display') else ''
        except Exception:
            statut_str = ''
        try:
            fond_initial_val = getattr(caisse, 'fond_initial', 0)
            fond_initial_str = f"{float(fond_initial_val):,.0f} FCFA"
        except Exception:
            fond_initial_str = ""

        caisse_info = [
            ["Nom de l'association:", caisse_nom],
            ["Code de la caisse:", caisse_code],
            ["Date de création:", caisse_date_str],
            ["Statut:", statut_str],
            ["Fonds initial:", fond_initial_str],
        ]
        
        # Ajouter la localisation si disponible
        try:
            if getattr(caisse, 'village', None):
                localisation = f"{caisse.village.nom}, {caisse.canton.nom}, {caisse.commune.nom}, {caisse.prefecture.nom}, {caisse.region.nom}"
                caisse_info.append(["Localisation:", localisation])
        except Exception:
            pass
        
        caisse_table = Table(caisse_info, colWidths=[2*inch, 4*inch])
        caisse_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(caisse_table)
        story.append(Spacer(1, 30))
        
        # Informations des responsables
        story.append(Paragraph("COMPTES UTILISATEURS DES RESPONSABLES", subtitle_style))
        
        for i, user_info in enumerate(created_users, 1):
            # Titre du responsable
            role_mapping = {
                'PRESIDENTE': 'PRÉSIDENTE',
                'SECRETAIRE': 'SECRÉTAIRE', 
                'TRESORIERE': 'TRÉSORIÈRE'
            }
            role_display = role_mapping.get(user_info['role'], user_info['role'])
            story.append(Paragraph(f"{i}. {role_display}", bold_style))
            
            # Informations du compte
            user_data = [
                ["Nom complet:", user_info['user'].get_full_name()],
                ["Nom d'utilisateur:", user_info['username']],
                ["Mot de passe:", user_info['password']],
                ["Rôle:", role_display],
                ["Date de création:", datetime.now().strftime('%d/%m/%Y à %H:%M')]
            ]
            
            user_table = Table(user_data, colWidths=[1.5*inch, 4.5*inch])
            user_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.lightblue),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTNAME', (0, 1), (1, 1), 'Helvetica-Bold'),  # Username en gras
                ('FONTNAME', (0, 2), (1, 2), 'Helvetica-Bold'),  # Password en gras
            ]))
            
            story.append(user_table)
            story.append(Spacer(1, 20))
        
        # Instructions importantes
        story.append(Spacer(1, 20))
        story.append(Paragraph("INSTRUCTIONS IMPORTANTES", subtitle_style))
        
        instructions = [
            "• Ce document contient les informations de connexion des responsables de la caisse.",
            "• Les mots de passe ont été générés automatiquement et sont sécurisés.",
            "• Communiquez ces informations aux responsables concernés de manière sécurisée.",
            "• Les responsables doivent changer leur mot de passe lors de leur première connexion.",
            "• Conservez ce document en lieu sûr pour référence future.",
            "• En cas de perte des identifiants, contactez l'administrateur du système."
        ]
        
        for instruction in instructions:
            story.append(Paragraph(f"  {instruction}", normal_style))
        
        story.append(Spacer(1, 30))
        
                # Pied de page
        footer_text = f"Document généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')} par l'administrateur du système"
        story.append(Paragraph(footer_text, ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            alignment=TA_CENTER,
            textColor=colors.grey
        )))
        
        # Construire le PDF
        doc.build(story)
        
        # Récupérer le contenu du buffer
        pdf_content = buffer.getvalue()
        buffer.close()
        
        return pdf_content
        
    except Exception as e:
        logger.error(f"Erreur lors de la génération du PDF des identifiants: {str(e)}")
        # Retourner un PDF d'erreur simple
        return generate_simple_error_pdf(caisse, created_users, str(e))


def generate_simple_error_pdf_caisse(caisse, error_message):
    """Génère un PDF simple en cas d'erreur pour une caisse"""
    try:
        buffer = BytesIO()
        c = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4
        y = height - 50

        c.setFont('Helvetica-Bold', 16)
        c.drawCentredString(width/2, y, 'LISTE DES MEMBRES')
        y -= 40

        c.setFont('Helvetica', 12)
        c.drawString(60, y, f"Caisse: {caisse.nom_association}")
        y -= 20
        c.drawString(60, y, f"Code: {caisse.code}")
        y -= 30

        c.setFont('Helvetica-Bold', 12)
        c.drawString(60, y, "Erreur lors de la génération du PDF:")
        y -= 20
        
        c.setFont('Helvetica', 10)
        c.drawString(80, y, error_message)
        y -= 30

        c.setFont('Helvetica', 8)
        c.drawString(60, y, f"Document généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}")

        c.showPage()
        c.save()
        pdf = buffer.getvalue()
        buffer.close()
        return pdf
        
    except Exception as e:
        logger.error(f"Erreur lors de la génération du PDF simple pour caisse: {str(e)}")
        # Retourner un PDF minimal
        buffer = BytesIO()
        c = canvas.Canvas(buffer, pagesize=A4)
        c.drawString(100, 750, "Erreur lors de la génération du PDF")
        c.drawString(100, 730, f"Erreur: {error_message}")
        c.showPage()
        c.save()
        pdf = buffer.getvalue()
        buffer.close()
        return pdf


def generate_simple_error_pdf_membre(membre, error_message):
    """Génère un PDF simple en cas d'erreur pour un membre"""
    try:
        buffer = BytesIO()
        c = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4
        y = height - 50

        c.setFont('Helvetica-Bold', 16)
        c.drawCentredString(width/2, y, 'FICHE MEMBRE')
        y -= 40

        c.setFont('Helvetica', 12)
        c.drawString(60, y, f"Membre: {membre.nom_complet}")
        y -= 20
        c.drawString(60, y, f"Caisse: {membre.caisse.nom_association if membre.caisse else 'Aucune'}")
        y -= 30

        c.setFont('Helvetica-Bold', 12)
        c.drawString(60, y, "Erreur lors de la génération du PDF:")
        y -= 20
        
        c.setFont('Helvetica', 10)
        c.drawString(80, y, error_message)
        y -= 30

        c.setFont('Helvetica', 8)
        c.drawString(60, y, f"Document généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}")

        c.showPage()
        c.save()
        pdf = buffer.getvalue()
        buffer.close()
        return pdf
        
    except Exception as e:
        logger.error(f"Erreur lors de la génération du PDF simple pour membre: {str(e)}")
        # Retourner un PDF minimal
        buffer = BytesIO()
        c = canvas.Canvas(buffer, pagesize=A4)
        c.drawString(100, 750, "Erreur lors de la génération du PDF")
        c.drawString(100, 730, f"Erreur: {error_message}")
        c.showPage()
        c.save()
        pdf = buffer.getvalue()
        buffer.close()
        return pdf


def generate_simple_error_pdf(caisse, created_users, error_message):
    """Génère un PDF simple en cas d'erreur"""
    try:
        buffer = BytesIO()
        c = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4
        y = height - 50

        c.setFont('Helvetica-Bold', 16)
        c.drawCentredString(width/2, y, 'INFORMATIONS DE CONNEXION')
        y -= 40

        c.setFont('Helvetica', 12)
        c.drawString(60, y, f"Caisse: {caisse.nom_association}")
        y -= 20
        c.drawString(60, y, f"Code: {caisse.code}")
        y -= 30

        for user_info in created_users:
            c.setFont('Helvetica-Bold', 12)
            c.drawString(60, y, f"{user_info['role']}:")
            y -= 20
            
            c.setFont('Helvetica', 10)
            c.drawString(80, y, f"Nom: {user_info['user'].get_full_name()}")
            y -= 15
            c.drawString(80, y, f"Utilisateur: {user_info['username']}")
            y -= 15
            c.drawString(80, y, f"Mot de passe: {user_info['password']}")
            y -= 25

        c.showPage()
        c.save()
        pdf = buffer.getvalue()
        buffer.close()
        return pdf
        
    except Exception as e:
        logger.error(f"Erreur lors de la génération du PDF simple: {str(e)}")
        # Retourner un PDF minimal
        buffer = BytesIO()
        c = canvas.Canvas(buffer, pagesize=A4)
        c.drawString(100, 750, "Erreur lors de la génération du PDF")
        c.drawString(100, 730, f"Erreur: {error_message}")
        c.showPage()
        c.save()
        pdf = buffer.getvalue()
        buffer.close()
        return pdf


def create_credentials_pdf_response(caisse, created_users):
    """
    Crée une réponse HTTP avec le PDF des identifiants
    """
    try:
        pdf_content = generate_credentials_pdf(caisse, created_users)
        
        # Créer la réponse HTTP
        response = HttpResponse(pdf_content, content_type='application/pdf')
        safe_code = getattr(caisse, 'code', 'nouvelle_caisse')
        response['Content-Disposition'] = f'attachment; filename="identifiants_caisse_{safe_code}_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf"'
        
        return response
        
    except Exception as e:
        logger.error(f"Erreur dans create_credentials_pdf_response: {str(e)}")
        # Retourner un PDF minimal au lieu d'un fichier texte
        try:
            buffer = BytesIO()
            c = canvas.Canvas(buffer, pagesize=A4)
            c.setFont("Helvetica-Bold", 14)
            c.drawString(50, 800, "Informations de connexion - Caisse")
            c.setFont("Helvetica", 11)
            c.drawString(50, 770, "Un incident est survenu lors de la génération détaillée du PDF.")
            c.drawString(50, 750, "Le document minimal ci-dessous contient tout de même les identifiants.")
            
            y = 720
            for idx, u in enumerate(created_users, start=1):
                c.setFont("Helvetica-Bold", 12)
                c.drawString(50, y, f"{idx}. Rôle: {u.get('role', '')}")
                y -= 18
                c.setFont("Helvetica", 11)
                c.drawString(70, y, f"Nom d'utilisateur: {u.get('username', '')}")
                y -= 16
                c.drawString(70, y, f"Mot de passe: {u.get('password', '')}")
                y -= 22
                if y < 100:
                    c.showPage()
                    y = 800
            
            c.setFont("Helvetica-Oblique", 9)
            c.drawString(50, 60, f"Note: {str(e)}")
            c.showPage()
            c.save()
            pdf = buffer.getvalue()
            buffer.close()
            response = HttpResponse(pdf, content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="identifiants_caisse_minimal.pdf"'
            return response
        except Exception:
            # En tout dernier recours, renvoyer du texte
            error_response = HttpResponse(
                f"Erreur lors de la génération du PDF: {str(e)}", 
                content_type='text/plain'
            )
            error_response['Content-Disposition'] = 'attachment; filename="erreur_pdf.txt"'
            return error_response


def generate_pret_octroi_pdf(pret, buffer=None):
    """Génère un PDF moderne d'attestation d'octroi de prêt."""
    if buffer is None:
        buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    story = []
    
    # Récupérer les paramètres de l'application
    parametres = get_parametres_application()
    
    # Styles modernes
    styles = getSampleStyleSheet()
    
    # Style de section
    section_style = ParagraphStyle(
        'Section',
        parent=styles['Heading3'],
        fontSize=14,
        spaceAfter=15,
        spaceBefore=20,
        textColor=colors.HexColor('#F18F01'),
        fontName='Helvetica-Bold'
    )
    
    # En-tête standard avec logo, nom de l'application et informations
    create_standard_header(story, parametres, "ATTESTATION D'OCTROI DE PRÊT", f"Prêt N°: {pret.numero_pret}")
    
    # Informations du prêt
    story.append(Paragraph("📋 INFORMATIONS DU PRÊT", section_style))
    
    pret_info = [
        ["Numéro de prêt:", pret.numero_pret],
        ["Membre bénéficiaire:", pret.membre.nom_complet],
        ["Numéro de carte d'électeur:", pret.membre.numero_carte_electeur],
        ["Caisse:", pret.caisse.nom_association],
        ["Date de demande:", pret.date_demande.strftime('%d/%m/%Y')],
        ["Date d'octroi:", pret.date_decaissement.strftime('%d/%m/%Y %H:%M') if pret.date_decaissement else 'N/A'],
        ["Montant demandé:", f"{pret.montant_demande:,.0f} FCFA"],
        ["Montant accordé:", f"{pret.montant_accord:,.0f} FCFA"],
        ["Taux d'intérêt:", f"{pret.taux_interet}%" if pret.taux_interet else "0%"],
        ["Durée:", f"{pret.duree_mois} mois"],
        ["Motif:", pret.motif or "Non spécifié"]
    ]
    
    pret_table = Table(pret_info, colWidths=[2.5*inch, 4*inch])
    pret_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E8F4FD')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#CCCCCC')),
        ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')])
    ]))
    
    story.append(pret_table)
    story.append(Spacer(1, 20))
    
    # Résumé financier
    story.append(Paragraph("💰 RÉSUMÉ FINANCIER", section_style))
    
    montant_principal = pret.montant_accord
    montant_interet = pret.montant_interet_calcule
    total_a_rembourser = pret.total_a_rembourser
    
    resume_financier = [
        ["Montant principal accordé:", f"{montant_principal:,.0f} FCFA"],
        ["Intérêts calculés:", f"{montant_interet:,.0f} FCFA"],
        ["Net à payer (après taux d'intérêt):", f"{total_a_rembourser:,.0f} FCFA"],
        ["Échéance mensuelle:", f"{(total_a_rembourser / pret.duree_mois):,.0f} FCFA"],
        ["Statut:", "✅ PRÊT OCTROYÉ"]
    ]
    
    resume_table = Table(resume_financier, colWidths=[2.5*inch, 4*inch])
    resume_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E8F5E8')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#CCCCCC')),
        ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')])
    ]))
    
    story.append(resume_table)
    story.append(Spacer(1, 20))
    
    # Échéances de remboursement
    story.append(Paragraph("📅 ÉCHÉANCES DE REMBOURSEMENT", section_style))
    
    # Récupérer les échéances calculées
    echeances = pret.echeances.all().order_by('numero_echeance')
    
    if echeances.exists():
        echeances_headers = ["N° Échéance", "Date d'échéance", "Montant à payer", "Statut"]
        echeances_data = [echeances_headers]
        
        for echeance in echeances:
            echeances_data.append([
                f"Échéance {echeance.numero_echeance}",
                echeance.date_echeance.strftime('%d/%m/%Y'),
                f"{echeance.montant_echeance:,.0f} FCFA",
                echeance.get_statut_display()
            ])
        
        echeances_table = Table(echeances_data, colWidths=[1.5*inch, 1.5*inch, 1.5*inch, 1.5*inch])
        echeances_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#D4EDDA')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#28A745')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')])
        ]))
        
        story.append(echeances_table)
    else:
        # Si pas d'échéances calculées, afficher un message
        story.append(Paragraph(
            "⚠️ Les échéances de remboursement seront calculées automatiquement lors de l'octroi du prêt.",
            ParagraphStyle(
                'Warning',
                parent=styles['Normal'],
                fontSize=10,
                textColor=colors.HexColor('#856404'),
                fontName='Helvetica'
            )
        ))
    
    story.append(Spacer(1, 20))
    
    # Informations de la caisse
    story.append(Paragraph("🏛️ INFORMATIONS DE LA CAISSE", section_style))
    
    caisse_info = [
        ["Nom de l'association:", pret.caisse.nom_association],
        ["Code de la caisse:", pret.caisse.code],
        ["Région:", pret.caisse.region.nom if pret.caisse.region else "Non définie"],
        ["Préfecture:", pret.caisse.prefecture.nom if pret.caisse.prefecture else "Non définie"],
        ["Commune:", pret.caisse.commune.nom if pret.caisse.commune else "Non définie"],
        ["Présidente:", pret.caisse.presidente.nom_complet if pret.caisse.presidente else "Non définie"],
        ["Trésorière:", pret.caisse.tresoriere.nom_complet if pret.caisse.tresoriere else "Non définie"]
    ]
    
    caisse_table = Table(caisse_info, colWidths=[2.5*inch, 4*inch])
    caisse_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#FFF3E0')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#CCCCCC')),
        ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')])
    ]))
    
    story.append(caisse_table)
    story.append(Spacer(1, 20))
    
    # Conditions et obligations
    story.append(Paragraph("📋 CONDITIONS ET OBLIGATIONS", section_style))
    
    conditions_text = f"""
    <b>Le membre bénéficiaire s'engage à :</b><br/>
    • Rembourser le montant total de <b>{pret.total_a_rembourser:,.0f} FCFA</b> sur une durée de <b>{pret.duree_mois} mois</b><br/>
    • Respecter les échéances de remboursement mensuelles de <b>{(pret.total_a_rembourser / pret.duree_mois):,.0f} FCFA</b><br/>
    • Payer les intérêts de <b>{pret.taux_interet}%</b> inclus dans le total à rembourser<br/>
    • Informer la caisse en cas de difficultés de remboursement<br/>
    • Participer aux réunions de la caisse<br/><br/>
    
    <b>La caisse s'engage à :</b><br/>
    • Accompagner le membre dans son projet<br/>
    • Fournir un suivi régulier du remboursement<br/>
    • Respecter la confidentialité des informations<br/>
    • Apporter un soutien en cas de difficultés
    """
    
    story.append(Paragraph(conditions_text, ParagraphStyle(
        'Conditions',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=20,
        leftIndent=20,
        rightIndent=20,
        fontName='Helvetica'
    )))
    
    # Signature et validation
    story.append(Paragraph("✍️ SIGNATURES ET VALIDATION", section_style))
    
    # Récupérer les informations du Président Général depuis les paramètres
    nom_pg, titre_pg, sig_pg, _ = get_signature_president_general()
    
    # Créer le tableau des signatures avec tous les responsables requis
    signatures_data = []
    
    # Signature du Président Général
    signatures_data.append([
        f"{titre_pg} de toutes les caisses:",
        sig_pg,
        nom_pg
    ])
    
    # Signatures des responsables de la caisse
    if pret.caisse:
        # Présidente
        if pret.caisse.presidente:
            if validate_image_file(pret.caisse.presidente.signature):
                try:
                    sig_pres = Image(pret.caisse.presidente.signature.path, width=1*inch, height=0.5*inch)
                except Exception as e:
                    logger.warning(f"Erreur lors du chargement de la signature de la présidente: {e}")
                    sig_pres = ""
            else:
                sig_pres = ""
            
            signatures_data.append([
                "Présidente de la caisse:",
                sig_pres,
                pret.caisse.presidente.nom_complet
            ])
        else:
            signatures_data.append([
                "Présidente de la caisse:",
                "Non définie",
                "Non définie"
            ])
        
        # Trésorière
        if pret.caisse.tresoriere:
            if validate_image_file(pret.caisse.tresoriere.signature):
                try:
                    sig_tres = Image(pret.caisse.tresoriere.signature.path, width=1*inch, height=0.5*inch)
                except Exception as e:
                    logger.warning(f"Erreur lors du chargement de la signature de la trésorière: {e}")
                    sig_tres = ""
            else:
                sig_tres = ""
            
            signatures_data.append([
                "Trésorière de la caisse:",
                sig_tres,
                pret.caisse.tresoriere.nom_complet
            ])
        else:
            signatures_data.append([
                "Trésorière de la caisse:",
                "Non définie",
                "Non définie"
            ])
        
        # Secrétaire
        if pret.caisse.secretaire:
            if validate_image_file(pret.caisse.secretaire.signature):
                try:
                    sig_sec = Image(pret.caisse.secretaire.signature.path, width=1*inch, height=0.5*inch)
                except Exception as e:
                    logger.warning(f"Erreur lors du chargement de la signature de la secrétaire: {e}")
                    sig_sec = ""
            else:
                sig_sec = ""
            
            signatures_data.append([
                "Secrétaire de la caisse:",
                sig_sec,
                pret.caisse.secretaire.nom_complet
            ])
        else:
            signatures_data.append([
                "Secrétaire de la caisse:",
                "Non définie",
                "Non définie"
            ])
    
    # Réorganiser les signatures avec le demandeur en premier
    signatures_data = create_signatures_table_with_demandeur_first(pret, signatures_data)
    
    # Créer le tableau des signatures
    signatures_table = Table(signatures_data, colWidths=[2.5*inch, 2*inch, 2*inch])
    signatures_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E8F4FD')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'CENTER'),
        ('ALIGN', (2, 0), (2, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#2E86AB')),
        ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')])
    ]))
    
    story.append(signatures_table)
    story.append(Spacer(1, 20))
    
    # Informations importantes
    story.append(Paragraph("⚠️ INFORMATIONS IMPORTANTES", section_style))
    
    important_text = f"""
    <b>Ce document est officiel et certifie l'octroi du prêt.</b><br/><br/>
    
    • Le prêt a été octroyé le <b>{pret.date_decaissement.strftime('%d/%m/%Y à %H:%M') if pret.date_decaissement else 'N/A'}</b><br/>
    • Le montant de <b>{pret.montant_accord:,.0f} FCFA</b> a été décaissé de la caisse<br/>
    • Le total à rembourser est de <b>{pret.total_a_rembourser:,.0f} FCFA</b> (principal + intérêts)<br/>
    • Le statut du prêt est maintenant <b>"En cours"</b><br/>
    • Les remboursements doivent commencer selon le calendrier établi<br/>
    • En cas de retard, des pénalités peuvent s'appliquer<br/><br/>
    
    <b>Contact de la caisse :</b><br/>
    Téléphone : {pret.caisse.presidente.numero_telephone if pret.caisse.presidente else 'N/A'}<br/>
    Adresse : {pret.caisse.presidente.adresse if pret.caisse.presidente else 'N/A'}
    """
    
    story.append(Paragraph(important_text, ParagraphStyle(
        'Important',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=20,
        leftIndent=20,
        rightIndent=20,
        fontName='Helvetica'
    )))
    
    # Ajouter les informations de contact
    add_contact_info_to_pdf(story, parametres)
    
    # Pied de page standard avec informations du PDG
    create_standard_footer(story, parametres)
    
    # Construire le PDF
    doc.build(story)
    
    # Récupérer le contenu du buffer
    pdf_content = buffer.getvalue()
    buffer.close()
    
    return pdf_content


def generate_remboursement_pdf(pret, mouvement):
    """Génère un PDF moderne de reçu de remboursement."""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    story = []
    
    # Récupérer les paramètres de l'application
    parametres = get_parametres_application()
    
    # Styles modernes
    styles = getSampleStyleSheet()
    
    # Style de section
    section_style = ParagraphStyle(
        'Section',
        parent=styles['Heading3'],
        fontSize=14,
        spaceAfter=15,
        spaceBefore=20,
        textColor=colors.HexColor('#F18F01'),
        fontName='Helvetica-Bold'
    )
    
    # En-tête standard avec logo, nom de l'application et informations
    create_standard_header(story, parametres, "REÇU DE REMBOURSEMENT", f"Prêt N°: {pret.numero_pret}")
    
    # Informations du prêt
    story.append(Paragraph("📋 INFORMATIONS DU PRÊT", section_style))
    
    pret_info = [
        ["Numéro de prêt:", pret.numero_pret],
        ["Membre bénéficiaire:", pret.membre.nom_complet],
        ["Numéro de carte d'électeur:", pret.membre.numero_carte_electeur],
        ["Caisse:", pret.caisse.nom_association],
        ["Date de demande:", pret.date_demande.strftime('%d/%m/%Y')],
        ["Montant demandé:", f"{pret.montant_demande:,.0f} FCFA"],
        ["Montant accordé:", f"{pret.montant_accord:,.0f} FCFA"],
        ["Taux d'intérêt:", f"{pret.taux_interet}%" if pret.taux_interet else "0%"],
        ["Durée:", f"{pret.duree_mois} mois"],
        ["Motif:", pret.motif or "Non spécifié"]
    ]
    
    pret_table = Table(pret_info, colWidths=[2.5*inch, 4*inch])
    pret_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E8F4FD')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#CCCCCC')),
        ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')])
    ]))
    
    story.append(pret_table)
    story.append(Spacer(1, 20))
    
    # Détails du remboursement
    story.append(Paragraph("💰 DÉTAILS DU REMBOURSEMENT", section_style))
    
    # Montants liés au remboursement courant
    montant_rembourse = mouvement.montant
    interet_rembourse = getattr(mouvement, 'interet_rembourse', 0) or 0
    total_rembourse = montant_rembourse + interet_rembourse

    # Calculs globaux (alignés avec l'attestation de prêt)
    net_a_payer = pret.total_a_rembourser  # Montant accordé + intérêts
    montant_cumule = pret.montant_rembourse  # Cumul principal remboursé

    # Si les intérêts cumulés ne sont pas historisés sur les mouvements,
    # on considère 0 par défaut pour rester cohérent avec les données actuelles
    interets_cumules = 0

    # Reste à payer basé sur le Net à payer
    reste_a_payer = net_a_payer - (montant_cumule + interets_cumules)
    
    remboursement_info = [
        ["Date de remboursement:", mouvement.date_mouvement.strftime('%d/%m/%Y à %H:%M')],
        ["Montant principal remboursé:", f"{montant_rembourse:,.0f} FCFA"],
        ["Intérêts remboursés:", f"{interet_rembourse:,.0f} FCFA"],
        ["Net à payer (après taux d'intérêt):", f"{net_a_payer:,.0f} FCFA"],
        ["Total remboursé:", f"{total_rembourse:,.0f} FCFA"],
        ["Montant remboursé cumulé:", f"{montant_cumule:,.0f} FCFA"],
        ["Reste à payer:", f"{reste_a_payer:,.0f} FCFA"],
        ["Statut du prêt:", "✅ REMBOURSEMENT PARTIEL" if reste_a_payer > 0 else "✅ REMBOURSEMENT COMPLET"]
    ]
    
    remboursement_table = Table(remboursement_info, colWidths=[2.5*inch, 4*inch])
    remboursement_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#D4EDDA')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#28A745')),
        ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')])
    ]))
    
    story.append(remboursement_table)
    story.append(Spacer(1, 20))
    
    # État des échéances
    story.append(Paragraph("📅 ÉTAT DES ÉCHÉANCES", section_style))
    
    # Récupérer les échéances du prêt
    echeances = pret.echeances.all().order_by('numero_echeance')
    
    if echeances.exists():
        echeances_headers = ["N° Échéance", "Date d'échéance", "Montant", "Statut", "Date paiement"]
        echeances_data = [echeances_headers]
        
        for echeance in echeances:
            statut_color = {
                'A_PAYER': '🔴',
                'PARTIELLEMENT_PAYE': '🟡',
                'PAYE': '🟢',
                'EN_RETARD': '🔴'
            }.get(echeance.statut, '⚪')
            
            date_paiement = echeance.date_paiement.strftime('%d/%m/%Y') if echeance.date_paiement else 'Non payé'
            
            echeances_data.append([
                f"Échéance {echeance.numero_echeance}",
                echeance.date_echeance.strftime('%d/%m/%Y'),
                f"{echeance.montant_echeance:,.0f} FCFA",
                f"{statut_color} {echeance.get_statut_display()}",
                date_paiement
            ])
        
        echeances_table = Table(echeances_data, colWidths=[1.2*inch, 1.2*inch, 1.2*inch, 1.2*inch, 1.2*inch])
        echeances_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#D4EDDA')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#28A745')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')])
        ]))
        
        story.append(echeances_table)
        
        # Résumé des échéances
        echeances_payees = echeances.filter(statut='PAYE').count()
        echeances_en_retard = echeances.filter(statut='EN_RETARD').count()
        echeances_a_payer = echeances.filter(statut='A_PAYER').count()
        
        # Forcer la création d'échéances si manquantes pour les anciens prêts
        try:
            pret.get_or_create_echeances()
        except Exception:
            pass

        resume_echeances = [
            ["Échéances payées:", f"{echeances_payees}/{pret.duree_mois}"],
            ["Échéances en retard:", f"{echeances_en_retard}"],
            ["Échéances à payer:", f"{echeances_a_payer}"],
            ["Prochaine échéance:", pret.get_prochaine_echeance().date_echeance.strftime('%d/%m/%Y') if pret.get_prochaine_echeance() else "Aucune"]
        ]
        
        resume_table = Table(resume_echeances, colWidths=[2.5*inch, 2.5*inch])
        resume_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E8F4FD')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#CCCCCC')),
            ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')])
        ]))
        
        story.append(Spacer(1, 10))
        story.append(resume_table)
    else:
        story.append(Paragraph(
            "⚠️ Aucune échéance calculée pour ce prêt.",
            ParagraphStyle(
                'Warning',
                parent=styles['Normal'],
                fontSize=10,
                textColor=colors.HexColor('#856404'),
                fontName='Helvetica'
            )
        ))
    
    story.append(Spacer(1, 20))
    
    # Informations de la caisse
    story.append(Paragraph("🏛️ INFORMATIONS DE LA CAISSE", section_style))
    
    caisse_info = [
        ["Nom de l'association:", pret.caisse.nom_association],
        ["Code de la caisse:", pret.caisse.code],
        ["Localisation:", f"{pret.caisse.village.nom if pret.caisse.village else 'N/A'}, {pret.caisse.commune.nom if pret.caisse.commune else 'N/A'}, {pret.caisse.prefecture.nom if pret.caisse.prefecture else 'N/A'}"],
        ["Fonds disponibles:", f"{pret.caisse.fond_disponible:,.0f} FCFA"],
        ["Solde avant remboursement:", f"{mouvement.solde_avant:,.0f} FCFA"],
        ["Solde après remboursement:", f"{mouvement.solde_apres:,.0f} FCFA"]
    ]
    
    caisse_table = Table(caisse_info, colWidths=[2.5*inch, 4*inch])
    caisse_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#FFF3CD')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#FFC107')),
        ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')])
    ]))
    
    story.append(caisse_table)
    story.append(Spacer(1, 30))
    
    # Ajouter les signatures
    story.append(Paragraph("✍️ SIGNATURES", section_style))
    
    # Récupérer les informations du Président Général depuis les paramètres
    nom_pg, titre_pg, sig_pg, _ = get_signature_president_general()
    
    # Créer le tableau des signatures avec tous les responsables requis
    signatures_data = []
    
    # Signature du Président Général
    signatures_data.append([
        f"{titre_pg} de toutes les caisses:",
        sig_pg,
        nom_pg
    ])
    
    # Signatures des responsables de la caisse
    if pret.caisse:
        # Présidente
        if pret.caisse.presidente:
            if validate_image_file(pret.caisse.presidente.signature):
                try:
                    sig_pres = Image(pret.caisse.presidente.signature.path, width=1*inch, height=0.5*inch)
                except Exception as e:
                    logger.warning(f"Erreur lors du chargement de la signature de la présidente: {e}")
                    sig_pres = ""
            else:
                sig_pres = ""
            
            signatures_data.append([
                "Présidente de la caisse:",
                sig_pres,
                pret.caisse.presidente.nom_complet
            ])
        else:
            signatures_data.append([
                "Présidente de la caisse:",
                "Non définie",
                "Non définie"
            ])
        
        # Trésorière
        if pret.caisse.tresoriere:
            if validate_image_file(pret.caisse.tresoriere.signature):
                try:
                    sig_tres = Image(pret.caisse.tresoriere.signature.path, width=1*inch, height=0.5*inch)
                except Exception as e:
                    logger.warning(f"Erreur lors du chargement de la signature de la trésorière: {e}")
                    sig_tres = ""
            else:
                sig_tres = ""
            
            signatures_data.append([
                "Trésorière de la caisse:",
                sig_tres,
                pret.caisse.tresoriere.nom_complet
            ])
        else:
            signatures_data.append([
                "Trésorière de la caisse:",
                "Non définie",
                "Non définie"
            ])
        
        # Secrétaire
        if pret.caisse.secretaire:
            if validate_image_file(pret.caisse.secretaire.signature):
                try:
                    sig_sec = Image(pret.caisse.secretaire.signature.path, width=1*inch, height=0.5*inch)
                except Exception as e:
                    logger.warning(f"Erreur lors du chargement de la signature de la secrétaire: {e}")
                    sig_sec = ""
            else:
                sig_sec = ""
            
            signatures_data.append([
                "Secrétaire de la caisse:",
                sig_sec,
                pret.caisse.secretaire.nom_complet
            ])
        else:
            signatures_data.append([
                "Secrétaire de la caisse:",
                "Non définie",
                "Non définie"
            ])
    
    # Réorganiser les signatures avec le demandeur en premier
    signatures_data = create_signatures_table_with_demandeur_first(pret, signatures_data)
    
    # Créer le tableau des signatures
    signatures_table = Table(signatures_data, colWidths=[2.5*inch, 2*inch, 2*inch])
    signatures_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E8F4FD')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'CENTER'),
        ('ALIGN', (2, 0), (2, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#2E86AB')),
        ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')])
    ]))
    
    story.append(signatures_table)
    story.append(Spacer(1, 20))
    
    # Message de confirmation
    confirmation_text = f"""
    <b>CONFIRMATION DE REMBOURSEMENT</b><br/><br/>
    Nous confirmons la réception du remboursement de <b>{total_rembourse:,.0f} FCFA</b> 
    pour le prêt N° {pret.numero_pret} accordé à {pret.membre.nom_complet}.<br/><br/>
    
    Ce reçu atteste que le montant a été correctement enregistré et que le solde 
    du prêt a été mis à jour. Le membre a maintenant un reste à payer de 
    <b>{reste_a_payer:,.0f} FCFA</b>.
    """
    
    story.append(Paragraph(confirmation_text, ParagraphStyle(
        'Confirmation',
        parent=styles['Normal'],
        fontSize=11,
        spaceAfter=20,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#28A745'),
        fontName='Helvetica-Bold'
    )))
    
    # Ajouter les informations de contact
    add_contact_info_to_pdf(story, parametres)
    
    # Pied de page standard avec informations du PDG
    create_standard_footer(story, parametres)
    
    # Construire le PDF
    doc.build(story)
    
    # Récupérer le contenu du buffer
    pdf_content = buffer.getvalue()
    buffer.close()
    
    return pdf_content


def generate_remboursement_complet_pdf(pret, mouvements_remboursement, buffer=None):
    """Génère un PDF moderne et complet de remboursement pour un prêt terminé."""
    if buffer is None:
        buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    story = []
    
    # Récupérer les paramètres de l'application
    parametres = get_parametres_application()
    
    # Styles modernes
    styles = getSampleStyleSheet()
    
    # Style de section
    section_style = ParagraphStyle(
        'Section',
        parent=styles['Heading3'],
        fontSize=14,
        spaceAfter=15,
        spaceBefore=20,
        textColor=colors.HexColor('#F18F01'),
        fontName='Helvetica-Bold'
    )
    
    # En-tête standard avec logo, nom de l'application et informations
    create_standard_header(story, parametres, "ATTESTATION DE REMBOURSEMENT COMPLET", f"Prêt N°: {pret.numero_pret}")
    
    # Informations du prêt
    story.append(Paragraph("📋 INFORMATIONS DU PRÊT", section_style))
    
    pret_info = [
        ["Numéro de prêt:", pret.numero_pret],
        ["Membre bénéficiaire:", pret.membre.nom_complet],
        ["Caisse:", pret.caisse.nom_association],
        ["Date de demande:", pret.date_demande.strftime('%d/%m/%Y')],
        ["Montant demandé:", f"{pret.montant_demande:,.0f} FCFA"],
        ["Taux d'intérêt:", f"{pret.taux_interet}%" if pret.taux_interet else "0%"],
        ["Montant accordé:", f"{pret.montant_accord:,.0f} FCFA"],
        ["Durée:", f"{pret.duree_mois} mois"],
        ["Motif:", pret.motif or "Non spécifié"]
    ]
    
    pret_table = Table(pret_info, colWidths=[2.5*inch, 4*inch])
    pret_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E8F4FD')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#CCCCCC')),
        ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')])
    ]))
    
    story.append(pret_table)
    story.append(Spacer(1, 20))
    
    # Résumé financier
    story.append(Paragraph("💰 RÉSUMÉ FINANCIER", section_style))
    
    # Total dû = principal + intérêts
    from decimal import Decimal
    montant_total = pret.total_a_rembourser or Decimal('0')
    montant_rembourse = pret.montant_rembourse  # cumul principal enregistré
    # On recalculera les intérêts payés ci-dessous à partir des mouvements
    interet_total = Decimal('0')
    
    resume_financier = [
        ["Montant total du prêt:", f"{montant_total:,.0f} FCFA"],
        ["Net à payer (après taux d'intérêt):", f"{pret.total_a_rembourser:,.0f} FCFA"],
        ["Total remboursé:", f"{montant_rembourse:,.0f} FCFA"],
        ["Intérêts payés:", f"{interet_total:,.0f} FCFA"],
        ["Statut:", "✅ REMBOURSEMENT COMPLET"]
    ]
    
    resume_table = Table(resume_financier, colWidths=[2.5*inch, 4*inch])
    resume_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#D4EDDA')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#28A745')),
        ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')])
    ]))
    
    story.append(resume_table)
    story.append(Spacer(1, 20))
    
    # Détail des remboursements
    story.append(Paragraph("📊 DÉTAIL DES REMBOURSEMENTS", section_style))
    
    # En-têtes du tableau des remboursements
    remboursement_headers = [
        "N°", "Date", "Montant principal", "Intérêts", "Total", "Solde restant"
    ]
    
    remboursement_data = [remboursement_headers]
    solde_cumule = Decimal('0')
    principal_restant = pret.montant_accord or Decimal('0')
    
    for i, mouvement in enumerate(sorted(mouvements_remboursement, key=lambda m: m.date_mouvement), 1):
        total_paye = Decimal(mouvement.montant)
        # Décomposer le paiement en part principale et intérêt
        principal_paye = total_paye if total_paye <= principal_restant else principal_restant
        interet = total_paye - principal_paye
        interet_total += interet
        solde_cumule += total_paye
        principal_restant -= principal_paye
        reste = montant_total - solde_cumule
        if reste < 0:
            reste = Decimal('0')
        
        remboursement_data.append([
            str(i),
            mouvement.date_mouvement.strftime('%d/%m/%Y'),
            f"{principal_paye:,.0f}",
            f"{interet:,.0f}",
            f"{total_paye:,.0f}",
            f"{reste:,.0f}"
        ])
    
    remboursement_table = Table(remboursement_data, colWidths=[0.5*inch, 1.2*inch, 1.5*inch, 1*inch, 1.2*inch, 1.2*inch])
    remboursement_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E86AB')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#CCCCCC')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')])
    ]))
    
    story.append(remboursement_table)
    story.append(Spacer(1, 20))
    
    # Ajouter les signatures
    story.append(Paragraph("✍️ SIGNATURES", section_style))
    
    # Récupérer les informations du Président Général depuis les paramètres
    nom_pg, titre_pg, sig_pg, _ = get_signature_president_general()
    
    # Créer le tableau des signatures avec tous les responsables requis
    signatures_data = []
    
    # Signature du Président Général
    signatures_data.append([
        f"{titre_pg} de toutes les caisses:",
        sig_pg,
        nom_pg
    ])
    
    # Signatures des responsables de la caisse
    if pret.caisse:
        # Présidente
        if pret.caisse.presidente:
            if validate_image_file(pret.caisse.presidente.signature):
                try:
                    sig_pres = Image(pret.caisse.presidente.signature.path, width=1*inch, height=0.5*inch)
                except Exception as e:
                    logger.warning(f"Erreur lors du chargement de la signature de la présidente: {e}")
                    sig_pres = ""
            else:
                sig_pres = ""
            
            signatures_data.append([
                "Présidente de la caisse:",
                sig_pres,
                pret.caisse.presidente.nom_complet
            ])
        else:
            signatures_data.append([
                "Présidente de la caisse:",
                "Non définie",
                "Non définie"
            ])
        
        # Trésorière
        if pret.caisse.tresoriere:
            if validate_image_file(pret.caisse.tresoriere.signature):
                try:
                    sig_tres = Image(pret.caisse.tresoriere.signature.path, width=1*inch, height=0.5*inch)
                except Exception as e:
                    logger.warning(f"Erreur lors du chargement de la signature de la trésorière: {e}")
                    sig_tres = ""
            else:
                sig_tres = ""
            
            signatures_data.append([
                "Trésorière de la caisse:",
                sig_tres,
                pret.caisse.tresoriere.nom_complet
            ])
        else:
            signatures_data.append([
                "Trésorière de la caisse:",
                "Non définie",
                "Non définie"
            ])
        
        # Secrétaire
        if pret.caisse.secretaire:
            if validate_image_file(pret.caisse.secretaire.signature):
                try:
                    sig_sec = Image(pret.caisse.secretaire.signature.path, width=1*inch, height=0.5*inch)
                except Exception as e:
                    logger.warning(f"Erreur lors du chargement de la signature de la secrétaire: {e}")
                    sig_sec = ""
            else:
                sig_sec = ""
            
            signatures_data.append([
                "Secrétaire de la caisse:",
                sig_sec,
                pret.caisse.secretaire.nom_complet
            ])
        else:
            signatures_data.append([
                "Secrétaire de la caisse:",
                "Non définie",
                "Non définie"
            ])
    
    # Réorganiser les signatures avec le demandeur en premier
    signatures_data = create_signatures_table_with_demandeur_first(pret, signatures_data)
    
    # Créer le tableau des signatures
    signatures_table = Table(signatures_data, colWidths=[2.5*inch, 2*inch, 2*inch])
    signatures_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E8F4FD')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'CENTER'),
        ('ALIGN', (2, 0), (2, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#2E86AB')),
        ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')])
    ]))
    
    story.append(signatures_table)
    story.append(Spacer(1, 20))
    
    # Ajouter les informations de contact
    add_contact_info_to_pdf(story, parametres)
    
    # Pied de page standard avec informations du PDG
    create_standard_footer(story, parametres)
    
    # Construire le PDF
    doc.build(story)
    
    # Récupérer le contenu du buffer
    pdf_content = buffer.getvalue()
    buffer.close()
    
    return pdf_content
    


def create_agent_credentials_pdf_response(agent, created_user):
    """
    Génère un PDF avec les informations de connexion d'un agent
    """
    try:
        # Créer le buffer pour le PDF
        buffer = BytesIO()
        
        # Créer le document PDF
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        story = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=TA_CENTER,
            textColor=colors.darkblue
        )
        
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Heading2'],
            fontSize=14,
            spaceAfter=20,
            alignment=TA_CENTER,
            textColor=colors.darkgreen
        )
        
        normal_style = styles['Normal']
        bold_style = ParagraphStyle(
            'Bold',
            parent=styles['Normal'],
            fontSize=12,
            spaceAfter=6,
            textColor=colors.black
        )
        
        # Titre principal
        story.append(Paragraph("IDENTIFIANTS DE CONNEXION - AGENT", title_style))
        story.append(Spacer(1, 20))
        
        # Informations de l'agent
        story.append(Paragraph("INFORMATIONS DE L'AGENT", subtitle_style))
        
        agent_info = [
            ["Nom complet:", agent.nom_complet],
            ["Matricule:", agent.matricule],
            ["Numéro de carte d'électeur:", agent.numero_carte_electeur],
            ["Date d'embauche:", agent.date_embauche.strftime('%d/%m/%Y')],
            ["Statut:", agent.get_statut_display()],
            ["Téléphone:", agent.numero_telephone],
            ["Email:", agent.email or "Non renseigné"],
        ]
        
        # Ajouter la zone de responsabilité si définie
        if agent.region:
            zone_resp = f"{agent.region.nom}"
            if agent.prefecture:
                zone_resp += f", {agent.prefecture.nom}"
            agent_info.append(["Zone de responsabilité:", zone_resp])
        
        agent_table = Table(agent_info, colWidths=[2*inch, 4*inch])
        agent_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(agent_table)
        story.append(Spacer(1, 30))
        
        # Informations de connexion
        story.append(Paragraph("COMPTE UTILISATEUR", subtitle_style))
        
        user_data = [
            ["Nom complet:", created_user['user'].get_full_name()],
            ["Nom d'utilisateur:", created_user['username']],
            ["Mot de passe:", created_user['password']],
            ["Rôle:", created_user['role']],
            ["Date de création:", datetime.now().strftime('%d/%m/%Y à %H:%M')]
        ]
        
        user_table = Table(user_data, colWidths=[1.5*inch, 4.5*inch])
        user_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.lightblue),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (1, 1), 'Helvetica-Bold'),  # Username en gras
            ('FONTNAME', (0, 2), (1, 2), 'Helvetica-Bold'),  # Password en gras
        ]))
        
        story.append(user_table)
        story.append(Spacer(1, 30))
        
        # Instructions importantes
        story.append(Paragraph("INSTRUCTIONS IMPORTANTES", subtitle_style))
        
        instructions = [
            "• Ce document contient les informations de connexion de l'agent.",
            "• Le mot de passe a été généré automatiquement et est sécurisé.",
            "• L'agent doit changer son mot de passe lors de sa première connexion.",
            "• L'agent aura accès uniquement aux caisses qui lui sont assignées.",
            "• Conservez ce document en lieu sûr pour référence future.",
            "• En cas de perte des identifiants, contactez l'administrateur du système.",
            "• L'agent peut consulter la liste de ses caisses depuis son tableau de bord."
        ]
        
        for instruction in instructions:
            story.append(Paragraph(f"  {instruction}", normal_style))
        
        story.append(Spacer(1, 30))
        
        # Pied de page
        footer_text = f"Document généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')} par l'administrateur du système"
        story.append(Paragraph(footer_text, ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            alignment=TA_CENTER,
            textColor=colors.grey
        )))
        
        # Construire le PDF
        doc.build(story)
        
        # Récupérer le contenu du buffer
        pdf_content = buffer.getvalue()
        buffer.close()
        
        # Retourner une réponse HTTP PDF (cohérent avec la version caisse)
        response = HttpResponse(pdf_content, content_type='application/pdf')
        safe_id = getattr(agent, 'matricule', None) or getattr(agent, 'id', 'agent')
        response['Content-Disposition'] = f'attachment; filename="identifiants_agent_{safe_id}_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf"'
        return response
        
    except Exception as e:
        logger.error(f"Erreur lors de la génération du PDF agent: {str(e)}")
        # Retourner un PDF simple en cas d'erreur (enveloppe HttpResponse)
        pdf_fallback = generate_simple_agent_error_pdf(agent, created_user, str(e))
        response = HttpResponse(pdf_fallback, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="identifiants_agent_minimal.pdf"'
        return response


def generate_simple_agent_error_pdf(agent, created_user, error_message):
    """Génère un PDF simple en cas d'erreur pour les agents"""
    try:
        buffer = BytesIO()
        c = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4
        y = height - 50

        c.setFont('Helvetica-Bold', 16)
        c.drawCentredString(width/2, y, 'IDENTIFIANTS DE CONNEXION - AGENT')
        y -= 40

        c.setFont('Helvetica', 12)
        c.drawString(60, y, f"Agent: {agent.nom_complet}")
        y -= 20
        c.drawString(60, y, f"Matricule: {agent.matricule}")
        y -= 30

        c.setFont('Helvetica-Bold', 12)
        c.drawString(60, y, "Compte utilisateur:")
        y -= 20
        
        c.setFont('Helvetica', 10)
        c.drawString(80, y, f"Nom: {created_user['user'].get_full_name()}")
        y -= 15
        c.drawString(80, y, f"Utilisateur: {created_user['username']}")
        y -= 15
        c.drawString(80, y, f"Mot de passe: {created_user['password']}")
        y -= 25

        c.showPage()
        c.save()
        pdf = buffer.getvalue()
        buffer.close()
        return pdf
        
    except Exception as e:
        logger.error(f"Erreur lors de la génération du PDF simple agent: {str(e)}")
        # Retourner un PDF minimal
        buffer = BytesIO()
        c = canvas.Canvas(buffer, pagesize=A4)
        c.drawString(100, 750, "Erreur lors de la génération du PDF agent")
        c.drawString(100, 730, f"Erreur: {error_message}")
        c.showPage()
        c.save()
        pdf = buffer.getvalue()
        buffer.close()
        return pdf


def generate_membres_liste_pdf(caisse):
    """Génère un PDF moderne de la liste des membres d'une caisse."""
    try:
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        story = []
        
        # Récupérer les paramètres de l'application
        parametres = get_parametres_application()
    
        # Styles modernes
        styles = getSampleStyleSheet()
        
        # Style de section
        section_style = ParagraphStyle(
            'Section',
            parent=styles['Heading3'],
            fontSize=14,
            spaceAfter=15,
            spaceBefore=20,
            textColor=colors.HexColor('#F18F01'),
            fontName='Helvetica-Bold'
        )
        
        # En-tête standard avec logo, nom de l'application et informations
        create_standard_header(story, parametres, "LISTE DES MEMBRES", f"Caisse: {caisse.nom_association}")
        
        # Informations de la caisse
        story.append(Paragraph("📋 INFORMATIONS DE LA CAISSE", section_style))
    
        caisse_info = [
            ["Nom de l'association:", caisse.nom_association],
            ["Code de la caisse:", caisse.code],
            ["Localisation:", f"{caisse.village.nom if caisse.village else 'N/A'}, {caisse.commune.nom if caisse.commune else 'N/A'}, {caisse.prefecture.nom if caisse.prefecture else 'N/A'}"],
            ["Date de création:", caisse.date_creation.strftime('%d/%m/%Y')],
            ["Fonds disponibles:", f"{caisse.fond_disponible:,.0f} FCFA"],
            ["Nombre total de membres:", f"{caisse.membres.count()}"]
        ]
        
        caisse_table = Table(caisse_info, colWidths=[2.5*inch, 4*inch])
        caisse_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E8F4FD')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#CCCCCC')),
            ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')])
        ]))
        
        story.append(caisse_table)
        story.append(Spacer(1, 20))
        
        # Liste des membres
        story.append(Paragraph("👥 LISTE DES MEMBRES", section_style))
    
        # En-têtes du tableau des membres
        membres_headers = [
            "N°", "Nom complet", "Téléphone", "Rôle", "Statut", "Date d'adhésion", "Carte électeur"
        ]
        
        membres_data = [membres_headers]
    
        for i, membre in enumerate(caisse.membres.all().order_by('nom', 'prenoms'), 1):
            role_mapping = {
                'PRESIDENTE': 'Présidente',
                'SECRETAIRE': 'Secrétaire',
                'TRESORIERE': 'Trésorière',
                'MEMBRE': 'Membre'
            }
            
            statut_mapping = {
                'ACTIF': 'Actif',
                'INACTIF': 'Inactif'
            }
            
            membres_data.append([
                str(i),
                membre.nom_complet,
                membre.numero_telephone or '-',
                role_mapping.get(membre.role, membre.role),
                statut_mapping.get(membre.statut, membre.statut),
                membre.date_adhesion.strftime('%d/%m/%Y') if membre.date_adhesion else '-',
                membre.numero_carte_electeur or '-'
            ])
        
        membres_table = Table(membres_data, colWidths=[0.5*inch, 2*inch, 1.2*inch, 1*inch, 0.8*inch, 1.2*inch, 1.5*inch])
        membres_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E86AB')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#CCCCCC')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')])
        ]))
        
        story.append(membres_table)
        story.append(Spacer(1, 20))
        
        # Statistiques des membres
        story.append(Paragraph("📊 STATISTIQUES DES MEMBRES", section_style))
    
        membres_actifs = caisse.membres.filter(statut='ACTIF').count()
        membres_inactifs = caisse.membres.filter(statut='INACTIF').count()
        total_membres = caisse.membres.count()
        
        stats_info = [
            ["Total des membres:", f"{total_membres}"],
            ["Membres actifs:", f"{membres_actifs}"],
            ["Membres inactifs:", f"{membres_inactifs}"],
            ["Taux d'activité:", f"{(membres_actifs/total_membres*100):.1f}%" if total_membres > 0 else "0%"]
        ]
        
        stats_table = Table(stats_info, colWidths=[2.5*inch, 4*inch])
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#D4EDDA')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#28A745')),
            ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')])
        ]))
        
        story.append(stats_table)
        story.append(Spacer(1, 30))
        
        # Ajouter les signatures
        story.append(Paragraph("✍️ SIGNATURES", section_style))
    
        # Récupérer les informations du Président Général depuis les paramètres
        nom_pg, titre_pg, sig_pg, _ = get_signature_president_general()
        
        # Créer le tableau des signatures
        signatures_data = []
        
        # Signature du Président Général
        signatures_data.append([
            f"{titre_pg} de toutes les caisses:",
            sig_pg,
            nom_pg
        ])
    
        # Signatures des responsables de la caisse
        if caisse:
            # Présidente
            if caisse.presidente:
                if validate_image_file(caisse.presidente.signature):
                    try:
                        sig_pres = Image(caisse.presidente.signature.path, width=1*inch, height=0.5*inch)
                    except Exception as e:
                        logger.warning(f"Erreur lors du chargement de la signature de la présidente: {e}")
                        sig_pres = ""
                else:
                    sig_pres = ""
                
                signatures_data.append([
                    "Présidente de la caisse:",
                    sig_pres,
                    caisse.presidente.nom_complet
                ])
            else:
                signatures_data.append([
                    "Présidente de la caisse:",
                    "Non définie",
                    "Non définie"
                ])
            
            # Trésorière
            if caisse.tresoriere:
                if validate_image_file(caisse.tresoriere.signature):
                    try:
                        sig_tres = Image(caisse.tresoriere.signature.path, width=1*inch, height=0.5*inch)
                    except Exception as e:
                        logger.warning(f"Erreur lors du chargement de la signature de la trésorière: {e}")
                        sig_tres = ""
                else:
                    sig_tres = ""
                
                signatures_data.append([
                    "Trésorière de la caisse:",
                    sig_tres,
                    caisse.tresoriere.nom_complet
                ])
            else:
                signatures_data.append([
                    "Trésorière de la caisse:",
                    "Non définie",
                    "Non définie"
                ])
            
            # Secrétaire
            if caisse.secretaire:
                if validate_image_file(caisse.secretaire.signature):
                    try:
                        sig_sec = Image(caisse.secretaire.signature.path, width=1*inch, height=0.5*inch)
                    except Exception as e:
                        logger.warning(f"Erreur lors du chargement de la signature de la secrétaire: {e}")
                        sig_sec = ""
                else:
                    sig_sec = ""
                
                signatures_data.append([
                    "Secrétaire de la caisse:",
                    sig_sec,
                    caisse.secretaire.nom_complet
                ])
            else:
                signatures_data.append([
                    "Secrétaire de la caisse:",
                    "Non définie",
                    "Non définie"
                ])
        
        # Créer le tableau des signatures
        signatures_table = Table(signatures_data, colWidths=[2.5*inch, 2*inch, 2*inch])
        signatures_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E8F4FD')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'CENTER'),
            ('ALIGN', (2, 0), (2, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTNAME', (2, 0), (2, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#2E86AB')),
            ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')])
        ]))
        
        story.append(signatures_table)
        story.append(Spacer(1, 20))
        
        # Pied de page standard avec informations du PDG
        create_standard_footer(story, parametres)
        
        # Construire le PDF
        doc.build(story)
        
        # Récupérer le contenu du buffer
        pdf_content = buffer.getvalue()
        buffer.close()
        
        return pdf_content
        
    except Exception as e:
        logger.error(f"Erreur lors de la génération du PDF du membre {membre.id}: {str(e)}")
        # Retourner un PDF d'erreur simple
        return generate_simple_error_pdf_membre(membre, str(e))
        
    except Exception as e:
        logger.error(f"Erreur lors de la génération du PDF de la liste des membres de la caisse {caisse.id}: {str(e)}")
        # Retourner un PDF d'erreur simple
        return generate_simple_error_pdf_caisse(caisse, str(e))


def generate_membre_individual_pdf(membre):
    """Génère un PDF moderne pour un membre individuel avec cadre photo et signatures."""
    try:
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        story = []
        
        # Récupérer les paramètres de l'application
        parametres = get_parametres_application()
    
        # Styles modernes
        styles = getSampleStyleSheet()
        
        # Style de section
        section_style = ParagraphStyle(
            'Section',
            parent=styles['Heading3'],
            fontSize=14,
            spaceAfter=15,
            spaceBefore=20,
            textColor=colors.HexColor('#F18F01'),
            fontName='Helvetica-Bold'
        )
        
        # En-tête standard avec logo, nom de l'application et informations
        create_standard_header(story, parametres, "FICHE MEMBRE", f"Membre: {membre.nom_complet}")
        
        # Informations du membre avec cadre photo
        story.append(Paragraph("👤 INFORMATIONS PERSONNELLES", section_style))
    
        # Créer un tableau avec photo et informations
        membre_info = [
            ["Photo:", "📷 CADRE PHOTO"],
            ["Nom complet:", membre.nom_complet],
            ["Numéro de carte d'électeur:", membre.numero_carte_electeur or 'Non renseigné'],
            ["Date de naissance:", membre.date_naissance.strftime('%d/%m/%Y') if membre.date_naissance else 'Non renseignée'],
            ["Sexe:", membre.get_sexe_display() if membre.sexe else 'Non renseigné'],
            ["Téléphone:", membre.numero_telephone or 'Non renseigné'],
            ["Adresse:", membre.adresse or 'Non renseignée']
        ]
        
        # Ajouter la photo du membre si elle existe
        if validate_image_file(membre.photo):
            try:
                # Créer un tableau avec la photo intégrée
                photo_cell = [[Image(membre.photo.path, width=1.5*inch, height=2*inch)]]
                photo_table = Table(photo_cell)
                photo_table.setStyle(TableStyle([
                    ('ALIGN', (0, 0), (0, 0), 'CENTER'),
                    ('VALIGN', (0, 0), (0, 0), 'MIDDLE'),
                    ('GRID', (0, 0), (0, 0), 1, colors.HexColor('#2E86AB')),
                    ('BACKGROUND', (0, 0), (0, 0), colors.HexColor('#F8F9FA')),
                ]))
                
                # Remplacer le placeholder par la vraie photo
                membre_info[0][1] = photo_table
            except Exception as e:
                # Si erreur avec la photo, garder le placeholder
                membre_info[0][1] = "📷 Photo non disponible"
                logger.warning(f"Erreur lors du chargement de la photo du membre {membre.id}: {e}")
        else:
            membre_info[0][1] = "📷 Aucune photo"
        
        membre_table = Table(membre_info, colWidths=[2.5*inch, 4*inch])
        membre_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E8F4FD')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#CCCCCC')),
            ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')])
        ]))
        
        story.append(membre_table)
        story.append(Spacer(1, 20))
        
        # Informations de la caisse
        story.append(Paragraph("🏛️ INFORMATIONS DE LA CAISSE", section_style))
    
        role_mapping = {
            'PRESIDENTE': 'Présidente',
            'SECRETAIRE': 'Secrétaire',
            'TRESORIERE': 'Trésorière',
            'MEMBRE': 'Membre'
        }
    
        statut_mapping = {
            'ACTIF': 'Actif',
            'INACTIF': 'Inactif'
        }
    
        caisse_info = [
            ["Caisse d'appartenance:", membre.caisse.nom_association if membre.caisse else 'Aucune'],
            ["Code de la caisse:", membre.caisse.code if membre.caisse else 'N/A'],
            ["Rôle dans la caisse:", role_mapping.get(membre.role, membre.role)],
            ["Statut:", statut_mapping.get(membre.statut, membre.statut)],
            ["Date d'adhésion:", membre.date_adhesion.strftime('%d/%m/%Y') if membre.date_adhesion else 'Non renseignée'],
            ["Localisation:", f"{membre.caisse.village.nom if membre.caisse and membre.caisse.village else 'N/A'}, {membre.caisse.commune.nom if membre.caisse and membre.caisse.commune else 'N/A'}, {membre.caisse.prefecture.nom if membre.caisse and membre.caisse.prefecture else 'N/A'}"]
        ]
    
        caisse_table = Table(caisse_info, colWidths=[2.5*inch, 4*inch])
        caisse_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#FFF3CD')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#FFC107')),
            ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')])
        ]))
        
        story.append(caisse_table)
        story.append(Spacer(1, 20))
    
        # Informations sur les prêts (si applicable)
        if membre.caisse:
            prets = membre.prets.all()
            if prets.exists():
                story.append(Paragraph("💰 HISTORIQUE DES PRÊTS", section_style))
            
                prets_headers = ["N° Prêt", "Montant", "Statut", "Date demande", "Reste à payer"]
                prets_data = [prets_headers]
                
                for pret in prets:
                    prets_data.append([
                        pret.numero_pret,
                        f"{pret.montant_accord:,.0f} FCFA" if pret.montant_accord else f"{pret.montant_demande:,.0f} FCFA",
                        pret.statut,
                        pret.date_demande.strftime('%d/%m/%Y'),
                        f"{pret.montant_restant:,.0f} FCFA" if pret.montant_restant else '0 FCFA'
                    ])
                
                prets_table = Table(prets_data, colWidths=[1.2*inch, 1.5*inch, 1*inch, 1.2*inch, 1.5*inch])
                prets_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#D4EDDA')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                    ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#28A745')),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')])
                ]))
                
                story.append(prets_table)
                story.append(Spacer(1, 20))
    
        # Message de confirmation
        confirmation_text = f"""
        <b>FICHE MEMBRE OFFICIELLE</b><br/><br/>
        Cette fiche atteste que {membre.nom_complet} est membre de la caisse {membre.caisse.nom_association if membre.caisse else 'N/A'} 
        avec le rôle de {role_mapping.get(membre.role, membre.role)}.<br/><br/>
        
        Le statut actuel du membre est <b>{statut_mapping.get(membre.statut, membre.statut)}</b> 
        et sa date d'adhésion est le <b>{membre.date_adhesion.strftime('%d/%m/%Y') if membre.date_adhesion else 'Non renseignée'}</b>.
        """
        
        story.append(Paragraph(confirmation_text, ParagraphStyle(
            'Confirmation',
            parent=styles['Normal'],
            fontSize=11,
            spaceAfter=20,
            alignment=TA_CENTER,
            textColor=colors.HexColor('#28A745'),
            fontName='Helvetica-Bold'
        )))
    
        # Ajouter les signatures
        story.append(Spacer(1, 30))
        story.append(Paragraph("✍️ SIGNATURES", section_style))
    
        # Créer le tableau des signatures
        signatures_data = []
    
        # Signature du Président Général/PDG depuis les paramètres
        nom_pg, titre_pg, sig_pg, _ = get_signature_president_general()
        signatures_data.append([
            f"{titre_pg}:",
            sig_pg,
            nom_pg
        ])
    
        # Signatures des responsables de la caisse
        if membre.caisse:
            # Présidente
            if membre.caisse.presidente:
                if validate_image_file(membre.caisse.presidente.signature):
                    try:
                        sig_pres = Image(membre.caisse.presidente.signature.path, width=1*inch, height=0.5*inch)
                    except Exception as e:
                        logger.warning(f"Erreur lors du chargement de la signature de la présidente: {e}")
                        sig_pres = ""
                else:
                    sig_pres = ""
                
                signatures_data.append([
                    "Présidente de la caisse:",
                    sig_pres,
                    membre.caisse.presidente.nom_complet
                ])
            else:
                signatures_data.append([
                    "Présidente de la caisse:",
                    "Non définie",
                    "Non définie"
                ])
        
            # Trésorière
            if membre.caisse.tresoriere:
                if validate_image_file(membre.caisse.tresoriere.signature):
                    try:
                        sig_tres = Image(membre.caisse.tresoriere.signature.path, width=1*inch, height=0.5*inch)
                    except Exception as e:
                        logger.warning(f"Erreur lors du chargement de la signature de la trésorière: {e}")
                        sig_tres = ""
                else:
                    sig_tres = ""
                
                signatures_data.append([
                    "Trésorière de la caisse:",
                    sig_tres,
                    membre.caisse.tresoriere.nom_complet
                ])
            else:
                signatures_data.append([
                    "Trésorière de la caisse:",
                    "Non définie",
                    "Non définie"
                ])
        
            # Secrétaire
            if membre.caisse.secretaire:
                if validate_image_file(membre.caisse.secretaire.signature):
                    try:
                        sig_sec = Image(membre.caisse.secretaire.signature.path, width=1*inch, height=0.5*inch)
                    except Exception as e:
                        logger.warning(f"Erreur lors du chargement de la signature de la secrétaire: {e}")
                        sig_sec = ""
                else:
                    sig_sec = ""
                
                signatures_data.append([
                    "Secrétaire de la caisse:",
                    sig_sec,
                    membre.caisse.secretaire.nom_complet
                ])
            else:
                signatures_data.append([
                    "Secrétaire de la caisse:",
                    "Non définie",
                    "Non définie"
                ])
    
        # Réorganiser les signatures avec le demandeur en premier
        signatures_data = create_signatures_table_with_demandeur_first(membre, signatures_data)
        
        # Créer le tableau des signatures
        signatures_table = Table(signatures_data, colWidths=[2.5*inch, 2*inch, 2*inch])
        signatures_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E8F4FD')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'CENTER'),
            ('ALIGN', (2, 0), (2, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTNAME', (2, 0), (2, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#2E86AB')),
            ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')])
        ]))
        
        story.append(signatures_table)
        story.append(Spacer(1, 20))
        
        # Pied de page standard avec informations du PDG
        create_standard_footer(story, parametres)
        
        # Construire le PDF
        doc.build(story)
        
        # Récupérer le contenu du buffer
        pdf_content = buffer.getvalue()
        buffer.close()
        
        return pdf_content
        
    except Exception as e:
        logger.error(f"Erreur lors de la génération du PDF du membre {membre.id}: {str(e)}")
        # Retourner un PDF d'erreur simple
        return generate_simple_error_pdf_membre(membre, str(e))


def _table_from_key_values(rows, col1_width, col2_width):
    table = RLTable(rows, colWidths=[col1_width, col2_width])
    table.setStyle(RLTableStyle([
         ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E8F4FD')),
         ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
         ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
         ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
         ('FONTSIZE', (0, 0), (-1, -1), 10),
         ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
         ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#CCCCCC')),
    ]))
    return table


def generate_rapport_pdf(rapport):
    """Génère un PDF structuré pour un RapportActivite."""
    buffer = BytesIO()
    # Marges réduites pour élargir les tableaux
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=18, rightMargin=18, topMargin=24, bottomMargin=24)
    story = []

    # Paramètres et en-tête standard
    parametres = get_parametres_application()
    # Styles nécessaires utilisés dans toutes les branches (y compris cotisations)
    styles = getSampleStyleSheet()
    titre_map = {
        'general': "RAPPORT GÉNÉRAL",
        'financier': "RAPPORT FINANCIER",
        'prets': "RAPPORT DES PRÊTS",
        'membres': "RAPPORT DES MEMBRES",
        'echeances': "RAPPORT DES ÉCHÉANCES",
    }
    # Ajouter libellés pour les rapports de cotisations afin d'éviter un double en-tête plus bas
    if getattr(rapport, 'type_rapport', '') in ('cotisations_general', 'cotisations_par_membre'):
        titre_map['cotisations_general'] = "RAPPORT COTISATIONS - GÉNÉRAL"
        titre_map['cotisations_par_membre'] = "RAPPORT COTISATIONS - PAR MEMBRE"
    sous_titre = rapport.caisse.nom_association if getattr(rapport, 'caisse', None) else "Toutes Caisses"
    if rapport.date_debut or rapport.date_fin:
        periode = f"Période: {rapport.date_debut.strftime('%d/%m/%Y') if rapport.date_debut else '-'} → {rapport.date_fin.strftime('%d/%m/%Y') if rapport.date_fin else '-'}"
    else:
        periode = "Période: Toutes"

    create_standard_header(story, parametres, titre_map.get(rapport.type_rapport, 'RAPPORT'), sous_titre)
    story.append(Paragraph(periode, ParagraphStyle('Periode', parent=getSampleStyleSheet()['Normal'], alignment=TA_CENTER)))
    story.append(Spacer(1, 10))

    data = rapport.donnees or {}

    # Extension: rapports cotisations (général, par membre agrégé, et membre individuel)
    if getattr(rapport, 'type_rapport', '') in ('cotisations_general', 'cotisations_par_membre', 'cotisations_membre'):
        items = rapport.donnees.get('items', []) if hasattr(rapport, 'donnees') else (rapport.items or [])
        totaux = rapport.donnees.get('totaux', {}) if hasattr(rapport, 'donnees') else {}

        # Construire le tableau
        rows = []
        if rapport.type_rapport == 'cotisations_general':
            # Retirer la colonne Fondation et la colonne Date
            rows.append(['Membre', 'Séance', 'Tempon', 'Solidarité', 'Pénalité', 'Total', 'Observation'])
            for it in items:
                rows.append([
                    it.get('membre',''),
                    it.get('seance',''),
                    f"{it.get('prix_tempon',0):,.0f}".replace(',', ' ').replace('.', ','),
                    f"{it.get('frais_solidarite',0):,.0f}".replace(',', ' ').replace('.', ','),
                    f"{it.get('penalite_emprunt_retard',0):,.0f}".replace(',', ' ').replace('.', ','),
                    f"{it.get('montant_total',0):,.0f}".replace(',', ' ').replace('.', ','),
                    (it.get('observation','') or '')[:80],
                ])
            # Largeurs compactes sans la colonne Date (somme ≈ 7.7")
            widths = [1.8*inch, 1.1*inch, 1.0*inch, 0.9*inch, 0.9*inch, 1.1*inch, 0.9*inch]
        elif rapport.type_rapport == 'cotisations_par_membre':
            # Retirer la colonne Fondation
            rows.append(['Membre', '#', 'Tempon', 'Solidarité', 'Pénalités', 'Total'])
            for it in items:
                rows.append([
                    it.get('membre',''),
                    str(it.get('nombre',0)),
                    f"{it.get('tempon',0):,.0f}".replace(',', ' ').replace('.', ','),
                    f"{it.get('solidarite',0):,.0f}".replace(',', ' ').replace('.', ','),
                    f"{it.get('penalite',0):,.0f}".replace(',', ' ').replace('.', ','),
                    f"{it.get('total',0):,.0f}".replace(',', ' ').replace('.', ','),
                ])
            # Largeurs compactes (somme ≈ 7.5")
            widths = [2.8*inch, 0.6*inch, 1.1*inch, 1.1*inch, 1.0*inch, 0.9*inch]
        else:
            # Rapport pour un seul membre: sans colonne Date, avec Séance + composantes
            membre = (rapport.donnees or {}).get('membre') if hasattr(rapport, 'donnees') else None
            rows.append(['Séance', 'Tempon', 'Solidarité', 'Pénalité', 'Total', 'Observation'])
            for it in items:
                rows.append([
                    it.get('seance',''),
                    f"{it.get('prix_tempon',0):,.0f}".replace(',', ' ').replace('.', ','),
                    f"{it.get('frais_solidarite',0):,.0f}".replace(',', ' ').replace('.', ','),
                    f"{it.get('penalite_emprunt_retard',0):,.0f}".replace(',', ' ').replace('.', ','),
                    f"{it.get('montant_total',0):,.0f}".replace(',', ' ').replace('.', ','),
                    (it.get('observation','') or '')[:80],
                ])
            widths = [1.2*inch, 1.0*inch, 1.0*inch, 1.0*inch, 1.2*inch, 2.0*inch]

        t = RLTable(rows, colWidths=widths, repeatRows=1)
        t.setStyle(RLTableStyle([
            ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#E8F4FD')),
            ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
            ('FONTSIZE',(0,0),(-1,0),9),
            ('GRID',(0,0),(-1,-1),0.5,colors.HexColor('#CCCCCC')),
            ('FONTSIZE',(0,1),(-1,-1),8),
            ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
            ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white, colors.HexColor('#F8F9FA')]),
            ('ALIGN',(-5,1),(-1,-1),'RIGHT'),
        ]))
        story.append(t)
        story.append(Spacer(1, 10))

        if totaux:
            # Afficher les totaux sans Fondation
            if rapport.type_rapport == 'cotisations_par_membre':
                story.append(Paragraph(
                    f"Totaux — Tempon: {totaux.get('tempon',0):,.0f}  Solidarité: {totaux.get('solidarite',0):,.0f}  Pénalités: {totaux.get('penalite',0):,.0f}  Total: {totaux.get('total',0):,.0f}  Nombre: {totaux.get('nombre',0):,.0f}".replace(',', ' ').replace('.', ','),
                    styles['Normal']
                ))
            else:
                story.append(Paragraph(
                    f"Totaux — Tempon: {totaux.get('tempon',0):,.0f}  Solidarité: {totaux.get('solidarite',0):,.0f}  Pénalités: {totaux.get('penalite',0):,.0f}  Total: {totaux.get('total',0):,.0f}".replace(',', ' ').replace('.', ','),
                    styles['Normal']
                ))

        # Ajouter un tableau de signatures esthétique: Trésorière, Secrétaire, Présidente et PCA
        try:
            caisse = getattr(rapport, 'caisse', None)
            nom_pg, titre_pg, sig_pg, _ = get_signature_president_general()
            # Préparer images et noms
            sig_pres = sig_sec = sig_tres = ""
            nom_pres = nom_sec = nom_tres = ""
            if caisse and getattr(caisse, 'presidente', None):
                nom_pres = getattr(caisse.presidente, 'nom_complet', str(caisse.presidente))
                if getattr(caisse.presidente, 'signature', None) and validate_image_file(caisse.presidente.signature):
                    sig_pres = Image(caisse.presidente.signature.path, width=1.4*inch, height=0.6*inch)
            if caisse and getattr(caisse, 'secretaire', None):
                nom_sec = getattr(caisse.secretaire, 'nom_complet', str(caisse.secretaire))
                if getattr(caisse.secretaire, 'signature', None) and validate_image_file(caisse.secretaire.signature):
                    sig_sec = Image(caisse.secretaire.signature.path, width=1.4*inch, height=0.6*inch)
            if caisse and getattr(caisse, 'tresoriere', None):
                nom_tres = getattr(caisse.tresoriere, 'nom_complet', str(caisse.tresoriere))
                if getattr(caisse.tresoriere, 'signature', None) and validate_image_file(caisse.tresoriere.signature):
                    sig_tres = Image(caisse.tresoriere.signature.path, width=1.4*inch, height=0.6*inch)

            # Construire le tableau: 4 colonnes, 3 lignes (titre, signature, nom)
            # Ordre: Trésorière, Secrétaire, Présidente, PCA
            headers_row = [
                Paragraph('<b>Trésorière</b>', styles['Normal']),
                Paragraph('<b>Secrétaire</b>', styles['Normal']),
                Paragraph('<b>Présidente</b>', styles['Normal']),
                Paragraph(f"<b>{titre_pg}</b>", styles['Normal'])
            ]
            images_row = [sig_tres or '', sig_sec or '', sig_pres or '', sig_pg or '']
            names_row = [
                Paragraph(nom_tres or '', styles['Normal']),
                Paragraph(nom_sec or '', styles['Normal']),
                Paragraph(nom_pres or '', styles['Normal']),
                Paragraph(nom_pg or '', styles['Normal'])
            ]

            sig_table = RLTable([headers_row, images_row, names_row], colWidths=[1.9*inch, 1.9*inch, 1.9*inch, 1.9*inch])
            sig_table.setStyle(RLTableStyle([
                ('GRID',(0,0),(-1,-1),0.25,colors.HexColor('#B0B0B0')),
                ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
                ('ALIGN',(0,0),(-1,-1),'CENTER'),
                ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#E8F4FD')),
                ('FONTSIZE',(0,0),(-1,-1),9),
                ('BOTTOMPADDING',(0,1),(-1,1),10),
                ('TOPPADDING',(0,1),(-1,1),10),
            ]))
            story.append(Spacer(1, 14))
            story.append(sig_table)
        except Exception:
            pass

        # Pied de page standard
        create_standard_footer(story, parametres)

        doc.build(story)
        pdf = buffer.getvalue()
        buffer.close()
        return pdf
    # Extension: rapport des dépenses (par période)
    elif getattr(rapport, 'type_rapport', '') == 'depenses':
        # Récupérer la caisse depuis le rapport
        caisse = getattr(rapport, 'caisse', None)
        
        items = (rapport.donnees or {}).get('items', []) if hasattr(rapport, 'donnees') else (rapport.items or [])
        totaux = (rapport.donnees or {}).get('totaux', {}) if hasattr(rapport, 'donnees') else {}

        rows = [['Date', 'Objectif', 'Montant', 'Observation']]
        for it in items:
            try:
                montant = float(it.get('montant', 0) or 0)
            except Exception:
                montant = 0
            rows.append([
                it.get('date', ''),
                (it.get('objectif', '') or '')[:80],
                f"{montant:,.0f}".replace(',', ' '),
                (it.get('observation', '') or '')[:80],
            ])

        t = RLTable(rows, colWidths=[1.2*inch, 3.4*inch, 1.0*inch, 1.9*inch], repeatRows=1)
        t.setStyle(RLTableStyle([
            ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#E8F4FD')),
            ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
            ('GRID',(0,0),(-1,-1),0.5,colors.HexColor('#CCCCCC')),
            ('FONTSIZE',(0,0),(-1,0),9),
            ('FONTSIZE',(0,1),(-1,-1),8),
            ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
            ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white, colors.HexColor('#F8F9FA')]),
            ('ALIGN',(2,1),(2,-1),'RIGHT'),
        ]))
        story.append(t)
        story.append(Spacer(1, 10))

        if totaux:
            try:
                somme = float(totaux.get('montant') or 0)
            except Exception:
                somme = 0
            story.append(Paragraph(
                f"Total des dépenses: {somme:,.0f} FCFA".replace(',', ' '),
                styles['Normal']
            ))
            
            # Afficher le solde restant dans la caisse
            if caisse:
                try:
                    # Vérifier si caisse est un objet modèle ou un dictionnaire
                    if hasattr(caisse, 'solde_disponible_depenses'):
                        solde_disponible = float(getattr(caisse, 'solde_disponible_depenses', 0) or 0)
                    else:
                        # Si c'est un dictionnaire, on ne peut pas calculer le solde
                        solde_disponible = 0
                    
                    if solde_disponible > 0:
                        story.append(Paragraph(
                            f"Solde disponible restant: {solde_disponible:,.0f} FCFA".replace(',', ' '),
                            styles['Normal']
                        ))
                except Exception:
                    pass

        # Signatures: Trésorière, Secrétaire, Présidente et PCA
        try:
            nom_pg, titre_pg, sig_pg, _ = get_signature_president_general()
            sig_pres = sig_sec = sig_tres = ""
            nom_pres = nom_sec = nom_tres = ""
            if caisse and hasattr(caisse, 'presidente') and getattr(caisse, 'presidente', None):
                nom_pres = getattr(caisse.presidente, 'nom_complet', str(caisse.presidente))
                if getattr(caisse.presidente, 'signature', None) and validate_image_file(caisse.presidente.signature):
                    sig_pres = Image(caisse.presidente.signature.path, width=1.4*inch, height=0.6*inch)
            if caisse and hasattr(caisse, 'secretaire') and getattr(caisse, 'secretaire', None):
                nom_sec = getattr(caisse.secretaire, 'nom_complet', str(caisse.secretaire))
                if getattr(caisse.secretaire, 'signature', None) and validate_image_file(caisse.secretaire.signature):
                    sig_sec = Image(caisse.secretaire.signature.path, width=1.4*inch, height=0.6*inch)
            if caisse and hasattr(caisse, 'tresoriere') and getattr(caisse, 'tresoriere', None):
                nom_tres = getattr(caisse.tresoriere, 'nom_complet', str(caisse.tresoriere))
                if getattr(caisse.tresoriere, 'signature', None) and validate_image_file(caisse.tresoriere.signature):
                    sig_tres = Image(caisse.tresoriere.signature.path, width=1.4*inch, height=0.6*inch)

            # Ordre: Trésorière, Secrétaire, Présidente, PCA
            headers_row = [
                Paragraph('<b>Trésorière</b>', styles['Normal']),
                Paragraph('<b>Secrétaire</b>', styles['Normal']),
                Paragraph('<b>Présidente</b>', styles['Normal']),
                Paragraph(f"<b>{titre_pg}</b>", styles['Normal'])
            ]
            images_row = [sig_tres or '', sig_sec or '', sig_pres or '', sig_pg or '']
            names_row = [
                Paragraph(nom_tres or '', styles['Normal']),
                Paragraph(nom_sec or '', styles['Normal']),
                Paragraph(nom_pres or '', styles['Normal']),
                Paragraph(nom_pg or '', styles['Normal'])
            ]
            sig_table = RLTable([headers_row, images_row, names_row], colWidths=[1.9*inch, 1.9*inch, 1.9*inch, 1.9*inch])
            sig_table.setStyle(RLTableStyle([
                ('GRID',(0,0),(-1,-1),0.25,colors.HexColor('#B0B0B0')),
                ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
                ('ALIGN',(0,0),(-1,-1),'CENTER'),
                ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#E8F4FD')),
                ('FONTSIZE',(0,0),(-1,-1),9),
                ('BOTTOMPADDING',(0,1),(-1,1),10),
                ('TOPPADDING',(0,1),(-1,1),10),
            ]))
            story.append(Spacer(1, 14))
            story.append(sig_table)
        except Exception:
            pass

        create_standard_footer(story, parametres)
        doc.build(story)
        pdf = buffer.getvalue()
        buffer.close()
        return pdf
    section_style = ParagraphStyle('Section', parent=styles['Heading3'], fontSize=13, spaceAfter=8, textColor=colors.HexColor('#2E86AB'))

    def add_dict_section(title, dct):
        story.append(Paragraph(title, section_style))
        rows = []
        for key, value in dct.items():
            rows.append([str(key).replace('_', ' ').title(), f"{value}"])
        story.append(_table_from_key_values(rows, 2.5*inch, 4*inch))
        story.append(Spacer(1, 10))

    # Helper pour graphique barre (utilisé par rapport financier)
    def add_bar_chart_from_stats(stats_list, title=""):
         try:
             # Ordonner par montant décroissant et limiter à un top N
             safe_stats = []
             for s in (stats_list or []):
                 try:
                     total = float(s.get('total') or 0)
                 except Exception:
                     total = 0
                 label_raw = (s.get('type_mouvement') or s.get('type') or '')
                 upper = (label_raw or '').upper()
                 label_norm = 'DECAISSEMENT' if upper == 'DECAISSEMENT' else ('REMBOURSEMENT' if upper == 'REMBOURSEMENT' else label_raw)
                 safe_stats.append({'label': label_norm, 'total': total})

             safe_stats.sort(key=lambda x: x['total'], reverse=True)

             TOP_N = 8
             top = safe_stats[:TOP_N]
             reste = safe_stats[TOP_N:]
             if reste:
                 total_autres = sum(x['total'] for x in reste)
                 top.append({'label': 'Autres', 'total': total_autres})

             # Construire labels abrégés et valeurs
             labels = []
             values = []
             for item in top:
                 lab = item['label'] or ''
                 # Abréviations compactes
                 up = lab.upper()
                 if up == 'DECAISSEMENT':
                     lab = 'Décaissement'
                 elif up == 'REMBOURSEMENT':
                     lab = 'Remboursement'
                 # Tronquer si trop long
                 if len(lab) > 12:
                     lab = lab[:11] + '…'
                 labels.append(lab)
                 values.append(float(item['total'] or 0))

             # Dessin
             dw = Drawing(460, 230)
             bc = VerticalBarChart()
             bc.x = 50
             bc.y = 40
             bc.height = 150
             bc.width = 360
             bc.data = [values]
             bc.categoryAxis.categoryNames = labels
             bc.valueAxis.valueMin = 0
             # Mise en forme des labels
             bc.categoryAxis.labels.angle = 45
             bc.categoryAxis.labels.fontSize = 7
             bc.valueAxis.labels.fontSize = 8
             bc.valueAxis.labelTextFormat = '{:,}'.format
             # Couleur
             bc.bars[0].fillColor = colors.HexColor('#2E86AB')
             # Espacement
             bc.barWidth = 12
             bc.groupSpacing = 6

             story.append(Paragraph(title or 'Graphique', styles['Heading4']))
             dw.add(bc)
             story.append(dw)
             story.append(Spacer(1, 10))
         except Exception as e:
             story.append(Paragraph(f"Graphique indisponible: {e}", styles['Normal']))

    def add_members_table(details, is_global=False):
         """Ajoute un tableau des membres. Si global, groupe par caisse."""
         try:
             if not details:
                 story.append(Paragraph("Aucun membre", styles['Normal']))
                 return
             if is_global:
                 # Grouper par caisse
                 groups = {}
                 for d in details:
                     caisse_nom = d.get('caisse') or 'Sans caisse'
                     groups.setdefault(caisse_nom, []).append(d)
                 for caisse_nom, rows in groups.items():
                     story.append(Spacer(1, 6))
                     story.append(Paragraph(f"Caisse: {caisse_nom}", styles['Heading4']))
                     data_rows = [[
                         'Nom', 'Carte', 'Rôle', 'Statut', 'Adhésion'
                     ]]
                     for m in rows:
                         data_rows.append([
                             f"{m.get('nom_complet','')}",
                             m.get('numero_carte','') or m.get('numero_carte_electeur',''),
                             m.get('role','') or '-',
                             m.get('statut','') or '-',
                             m.get('date_adhesion','') or '-',
                         ])
                     t = RLTable(data_rows, colWidths=[2.8*inch, 1.6*inch, 1.4*inch, 1.2*inch, 1.3*inch])
                     t.setStyle(RLTableStyle([
                         ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#E8F4FD')),
                         ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
                         ('FONTSIZE',(0,0),(-1,0),9),
                         ('FONTSIZE',(0,1),(-1,-1),8),
                         ('GRID',(0,0),(-1,-1),0.5,colors.HexColor('#CCCCCC')),
                         ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
                         ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white, colors.HexColor('#F8F9FA')]),
                     ]))
                     story.append(t)
             else:
                 # Caisse unique
                 data_rows = [[
                     'Nom', 'Carte', 'Rôle', 'Statut', 'Adhésion'
                 ]]
                 for m in details:
                     data_rows.append([
                         f"{m.get('nom_complet','')}",
                         m.get('numero_carte','') or m.get('numero_carte_electeur',''),
                         m.get('role','') or '-',
                         m.get('statut','') or '-',
                         m.get('date_adhesion','') or '-',
                     ])
                 t = RLTable(data_rows, colWidths=[3.0*inch, 1.7*inch, 1.4*inch, 1.2*inch, 1.3*inch])
                 t.setStyle(RLTableStyle([
                     ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#E8F4FD')),
                     ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
                     ('FONTSIZE',(0,0),(-1,0),9),
                     ('FONTSIZE',(0,1),(-1,-1),8),
                     ('GRID',(0,0),(-1,-1),0.5,colors.HexColor('#CCCCCC')),
                     ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
                     ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white, colors.HexColor('#F8F9FA')]),
                 ]))
                 story.append(t)
         except Exception as e:
             story.append(Paragraph(f"Erreur tableau membres: {e}", styles['Normal']))

    def add_stats_membres(stats_block):
         try:
             if not stats_block:
                 return
             # Par statut
             par_statut = stats_block.get('par_statut') or []
             if par_statut:
                 story.append(Paragraph('Statistiques des Membres', styles['Heading3']))
                 rows = [['Statut','Nombre']]
                 for s in par_statut:
                     rows.append([str(s.get('statut','-')), str(s.get('nombre',0))])
                 t = RLTable(rows, colWidths=[2.5*inch, 1.5*inch])
                 t.setStyle(RLTableStyle([
                     ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#E8F4FD')),
                     ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
                     ('GRID',(0,0),(-1,-1),0.5,colors.HexColor('#CCCCCC')),
                     ('FONTSIZE',(0,0),(-1,-1),9),
                 ]))
                 story.append(t)
                 story.append(Spacer(1,8))
             # Par sexe
             par_sexe = stats_block.get('par_sexe') or []
             if par_sexe:
                 rows = [['Sexe','Nombre']]
                 for s in par_sexe:
                     rows.append([str(s.get('sexe','-')), str(s.get('nombre',0))])
                 t = RLTable(rows, colWidths=[2.5*inch, 1.5*inch])
                 t.setStyle(RLTableStyle([
                     ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#E8F4FD')),
                     ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
                     ('GRID',(0,0),(-1,-1),0.5,colors.HexColor('#CCCCCC')),
                     ('FONTSIZE',(0,0),(-1,-1),9),
                 ]))
                 story.append(t)
                 story.append(Spacer(1,8))
             # Par rôle
             par_role = stats_block.get('par_role') or []
             if par_role:
                 rows = [['Rôle','Nombre']]
                 for s in par_role:
                     rows.append([str(s.get('role','-')), str(s.get('nombre',0))])
                 t = RLTable(rows, colWidths=[2.5*inch, 1.5*inch])
                 t.setStyle(RLTableStyle([
                     ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#E8F4FD')),
                     ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
                     ('GRID',(0,0),(-1,-1),0.5,colors.HexColor('#CCCCCC')),
                     ('FONTSIZE',(0,0),(-1,-1),9),
                 ]))
                 story.append(t)
         except Exception as e:
             story.append(Paragraph(f"Erreur statistiques membres: {e}", styles['Normal']))

    # Sections selon le type
    if rapport.type_rapport == 'general':
         if 'caisse' in data:
             add_dict_section('Caisse', data['caisse'])
         if 'membres' in data:
             add_dict_section('Membres', data['membres'])
         if 'prets' in data:
             add_dict_section('Prêts', data['prets'])
         if 'fonds' in data:
             add_dict_section('Fonds', data['fonds'])
    elif rapport.type_rapport == 'financier':
         if 'fonds_actuels' in data:
             add_dict_section('Fonds Actuels', data['fonds_actuels'])
         # Synthèse des prêts (par caisse)
         prets_syn = data.get('prets_synthese')
         if prets_syn:
             story.append(Paragraph('Synthèse des prêts', styles['Heading3']))
             rows = [
                 ['Total prêts (montant)', f"{float(prets_syn.get('total_prets_montant',0)) :,.0f}".replace(',', ' ') + ' FCFA'],
                 ['Total remboursé (montant)', f"{float(prets_syn.get('total_prets_rembourse_montant',0)) :,.0f}".replace(',', ' ') + ' FCFA'],
                 ['Reste à rembourser', f"{float(prets_syn.get('total_prets_restant_montant',0)) :,.0f}".replace(',', ' ') + ' FCFA'],
                 ['En attente (nombre)', str(prets_syn.get('nombre_en_attente',0))],
                 ['En cours (nombre)', str(prets_syn.get('nombre_en_cours',0))],
                 ['En retard (nombre)', str(prets_syn.get('nombre_en_retard',0))],
                 ['Remboursés (nombre)', str(prets_syn.get('nombre_rembourses',0))],
                 ['Taux de remboursement', f"{prets_syn.get('taux_remboursement',0)} %"],
                 ['Appréciation', prets_syn.get('appreciation','-')],
             ]
             t = RLTable(rows, colWidths=[2.8*inch, 3.2*inch])
             t.setStyle(RLTableStyle([
                 ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#E8F4FD')),
                 ('GRID',(0,0),(-1,-1),0.5,colors.HexColor('#CCCCCC')),
                 ('FONTSIZE',(0,0),(-1,-1),9),
             ]))
             story.append(t)
             story.append(Spacer(1,8))
         # Graphique des mouvements si dispo
         stats_mvts = (data.get('mouvements') or {}).get('stats_par_type')
         if stats_mvts:
             add_bar_chart_from_stats(stats_mvts, title="Mouvements par type")
         # Graphe prêts octroyés vs remboursés si dispo
         pf = data.get('prets_financiers') or {}
         if pf:
             try:
                 labels = ['Octroyés', 'Remboursés']
                 values = [float(pf.get('octroyes_total') or 0), float(pf.get('rembourses_total') or 0)]
                 dw = Drawing(400, 200)
                 bc = VerticalBarChart()
                 bc.x = 45; bc.y = 30; bc.height = 140; bc.width = 320
                 bc.data = [values]
                 bc.categoryAxis.categoryNames = labels
                 bc.valueAxis.valueMin = 0
                 bc.bars[0].fillColor = colors.HexColor('#6f42c1')
                 story.append(Paragraph('Prêts octroyés vs remboursés', styles['Heading4']))
                 dw.add(bc); story.append(dw); story.append(Spacer(1, 10))
             except Exception as e:
                 story.append(Paragraph(f"Graphique prêts indisponible: {e}", styles['Normal']))
         # Détail par caisse si global
         par_caisse = data.get('par_caisse')
         if par_caisse:
             story.append(Paragraph('Synthèse par caisse', styles['Heading3']))
             rows = [['Caisse', 'Fond initial', 'Fond dispo', 'Prêts total', 'Solde dispo']]
             for c in par_caisse:
                 rows.append([
                     f"{c.get('code','')} {c.get('nom','')}",
                     f"{c.get('fond_initial',0)}",
                     f"{c.get('fond_disponible',0)}",
                     f"{c.get('montant_total_prets',0)}",
                     f"{c.get('solde_disponible',0)}",
                 ])
             # Ligne de totaux
             try:
                 total_fi = sum(float(c.get('fond_initial') or 0) for c in par_caisse)
                 total_fd = sum(float(c.get('fond_disponible') or 0) for c in par_caisse)
                 total_prets = sum(float(c.get('montant_total_prets') or 0) for c in par_caisse)
                 total_solde = sum(float(c.get('solde_disponible') or 0) for c in par_caisse)
             except Exception:
                 total_fi = total_fd = total_prets = total_solde = 0
             rows.append(['TOTAL', f"{total_fi}", f"{total_fd}", f"{total_prets}", f"{total_solde}"])
             t = RLTable(rows, colWidths=[2.8*inch, 1.2*inch, 1.2*inch, 1.2*inch, 1.2*inch])
             t.setStyle(RLTableStyle([
                 ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#E8F4FD')),
                 ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
                 ('GRID',(0,0),(-1,-1),0.5,colors.HexColor('#CCCCCC')),
                 ('FONTSIZE',(0,0),(-1,0),9),
                 ('FONTSIZE',(0,1),(-1,-1),8),
                 ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white, colors.HexColor('#F8F9FA')]),
                 ('FONTNAME',(0,-1),(-1,-1),'Helvetica-Bold'),
                 ('BACKGROUND',(0,-1),(-1,-1),colors.HexColor('#F3F6FF')),
             ]))
             story.append(t)
         # Tableau des prêts par membre (états pertinents)
         prets_membres = data.get('prets_membres')
         if prets_membres:
             story.append(Spacer(1, 10))
             story.append(Paragraph('Prêts par membre', styles['Heading3']))
             # Enlever la colonne Caisse si le rapport est pour une seule caisse
             is_single_caisse = getattr(rapport, 'caisse', None) is not None
             if is_single_caisse:
                 headers = ['Membre', 'N° Prêt', 'Montant', 'Remboursé', 'Statut']
             else:
                 headers = ['Caisse', 'Membre', 'N° Prêt', 'Montant', 'Remboursé', 'Statut']
             rows = [headers]
             total_montant = 0.0
             total_rembourse = 0.0
             for r in prets_membres:
                 if is_single_caisse:
                     rows.append([
                         r.get('membre',''), r.get('numero_pret',''),
                         f"{float(r.get('montant_accord',0)) :,.0f}".replace(',', ' '),
                         f"{float(r.get('montant_rembourse',0)) :,.0f}".replace(',', ' '),
                         r.get('statut','')
                     ])
                 else:
                     caisse_label = r.get('caisse') or (getattr(rapport.caisse, 'nom', None) or '-')
                     rows.append([
                         caisse_label,
                         r.get('membre',''), r.get('numero_pret',''),
                         f"{float(r.get('montant_accord',0)) :,.0f}".replace(',', ' '),
                         f"{float(r.get('montant_rembourse',0)) :,.0f}".replace(',', ' '),
                         r.get('statut','')
                     ])
                 # Accumuler les totaux (utiliser valeurs brutes)
                 total_montant += float(r.get('montant_accord', 0) or 0)
                 total_rembourse += float(r.get('montant_rembourse', 0) or 0)
             # Ligne de totaux
             if is_single_caisse:
                 rows.append(['TOTAL', '', f"{total_montant:,.0f}".replace(',', ' '), f"{total_rembourse:,.0f}".replace(',', ' '), ''])
                 t = RLTable(rows, colWidths=[2.4*inch, 1.6*inch, 1.2*inch, 1.2*inch, 0.9*inch])
             else:
                 rows.append(['TOTAL', '', '', f"{total_montant:,.0f}".replace(',', ' '), f"{total_rembourse:,.0f}".replace(',', ' '), ''])
                 t = RLTable(rows, colWidths=[1.6*inch, 2.0*inch, 1.6*inch, 1.0*inch, 1.0*inch, 0.7*inch])
             styles_list = [
                 ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#E8F4FD')),
                 ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
                 ('GRID',(0,0),(-1,-1),0.5,colors.HexColor('#CCCCCC')),
                 ('FONTSIZE',(0,0),(-1,0),9),
                 ('FONTSIZE',(0,1),(-1,-1),8),
                 ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white, colors.HexColor('#F8F9FA')]),
             ]
             # Alignement des colonnes montants
             if is_single_caisse:
                 styles_list.append(('ALIGN',(2,1),(3,-2),'RIGHT'))
             else:
                 styles_list.append(('ALIGN',(3,1),(4,-2),'RIGHT'))
             # Mise en forme de la ligne TOTAL
             last_row = len(rows) - 1
             styles_list.extend([
                 ('FONTNAME',(0,last_row),(-1,last_row),'Helvetica-Bold'),
                 ('BACKGROUND',(0,last_row),(-1,last_row),colors.HexColor('#F3F6FF')),
                 ('ALIGN',(0,last_row),(-1,last_row),'RIGHT'),
             ])
             t.setStyle(RLTableStyle(styles_list))
             story.append(t)
    elif rapport.type_rapport == 'prets':
         stats_block = data.get('stats') or data.get('statistiques')
         if stats_block:
             # Tableau: Prêts par statut
             par_statut = stats_block.get('par_statut') or []
             if par_statut:
                 story.append(Paragraph('Statistiques des Prêts — par statut', styles['Heading3']))
                 rows = [['Statut', 'Nombre', 'Montant total']]
                 for s in par_statut:
                     montant_total = s.get('montant_total')
                     try:
                         montant_total = float(montant_total or 0)
                     except Exception:
                         montant_total = 0
                     rows.append([str(s.get('statut', '-')), str(s.get('nombre', 0)), f"{montant_total:,.0f}".replace(',', ' ')])
                 t = RLTable(rows, colWidths=[2.0*inch, 1.2*inch, 1.6*inch])
                 t.setStyle(RLTableStyle([
                     ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#E8F4FD')),
                     ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
                     ('GRID',(0,0),(-1,-1),0.5,colors.HexColor('#CCCCCC')),
                     ('ALIGN',(1,1),(2,-1),'RIGHT'),
                     ('FONTSIZE',(0,0),(-1,0),9),
                     ('FONTSIZE',(0,1),(-1,-1),8),
                 ]))
                 story.append(t)
                 story.append(Spacer(1, 8))

             # Tableau: Prêts par caisse (uniquement pour le global)
             par_caisse = stats_block.get('par_caisse') or []
             if par_caisse:
                 story.append(Paragraph('Statistiques des Prêts — par caisse', styles['Heading3']))
                 rows = [['Caisse', 'Nombre de prêts', 'Montant total']]
                 for s in par_caisse:
                     montant_total = s.get('montant_total') or s.get('montant_total_prets') or s.get('montant_total')
                     try:
                         montant_total = float(montant_total or 0)
                     except Exception:
                         montant_total = 0
                     rows.append([str(s.get('caisse__nom_association','-')), str(s.get('nombre_prets',0)), f"{montant_total:,.0f}".replace(',', ' ')])
                 t = RLTable(rows, colWidths=[3.2*inch, 1.4*inch, 1.6*inch])
                 t.setStyle(RLTableStyle([
                     ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#E8F4FD')),
                     ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
                     ('GRID',(0,0),(-1,-1),0.5,colors.HexColor('#CCCCCC')),
                     ('ALIGN',(1,1),(2,-1),'RIGHT'),
                     ('FONTSIZE',(0,0),(-1,0),9),
                     ('FONTSIZE',(0,1),(-1,-1),8),
                 ]))
                 story.append(t)
    elif rapport.type_rapport == 'membres':
         stats_block = data.get('stats') or data.get('statistiques')
         add_stats_membres(stats_block)
         # Détails des membres (global: par caisse, sinon liste de la caisse)
         details = data.get('details_membres')
         if details:
             story.append(Spacer(1, 10))
             story.append(Paragraph('Liste des membres', styles['Heading3']))
             add_members_table(details, is_global=(rapport.caisse is None))
    elif rapport.type_rapport == 'echeances':
         stats_block = data.get('stats') or data.get('statistiques')
         if stats_block:
             add_dict_section('Statistiques des Échéances', stats_block)

    # Pied de page
    create_standard_footer(story, parametres)
    doc.build(story)
    pdf_content = buffer.getvalue()
    buffer.close()
    return pdf_content


# Nouveaux rapports PDF système (global)
def generate_membres_systeme_pdf():
    """Génère un PDF listant tous les membres du système, groupés par caisse,
    en affichant l'état de la carte d'électeur (valide ou non)."""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=24, rightMargin=24, topMargin=28, bottomMargin=24)
    story = []

    parametres = get_parametres_application()
    create_standard_header(story, parametres, "MEMBRES - LISTE SYSTÈME", "Toutes Caisses")

    from .models import Caisse, Membre
    caisses = Caisse.objects.order_by('nom_association')

    styles = getSampleStyleSheet()
    section_title = ParagraphStyle('Sec', parent=styles['Heading3'], textColor=colors.HexColor('#2E86AB'))

    total_membres = 0
    total_valides = 0

    for caisse in caisses:
        story.append(Paragraph(f"Caisse: {caisse.nom_association}", section_title))
        membres = Membre.objects.filter(caisse=caisse).order_by('nom', 'prenoms')
        nb_caisse = membres.count()
        nb_valides = membres.filter(carte_electeur_valide=True).count()
        total_membres += nb_caisse
        total_valides += nb_valides
        story.append(Paragraph(f"Cartes valides: <b>{nb_valides}</b> / Membres: <b>{nb_caisse}</b>", styles['Normal']))
        story.append(Spacer(1, 6))
        rows = [[
            Paragraph('<b>Nom</b>', styles['Normal']),
            Paragraph('<b>N° Carte</b>', styles['Normal']),
            Paragraph('<b>Carte valide</b>', styles['Normal']),
            Paragraph('<b>Téléphone</b>', styles['Normal']),
        ]]
        for m in membres:
            etat = 'Oui' if getattr(m, 'carte_electeur_valide', False) else 'Non'
            rows.append([
                f"{m.nom} {m.prenoms}",
                m.numero_carte_electeur or '',
                etat,
                f"{m.indicatif_telephone or ''} {m.numero_telephone or ''}".strip(),
            ])

        table = RLTable(rows, colWidths=[2.6*inch, 1.7*inch, 1.3*inch, 1.7*inch])
        table.setStyle(RLTableStyle([
            ('GRID', (0,0), (-1,-1), 0.6, colors.HexColor('#CCCCCC')),
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#F0F6FF')),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('ALIGN', (2,1), (2,-1), 'CENTER'),
        ]))
        story.append(table)
        story.append(Spacer(1, 12))

    # Résumé global
    story.append(Spacer(1, 8))
    resume = RLTable([
        [Paragraph('<b>Total cartes valides</b>', styles['Normal']), str(total_valides)],
        [Paragraph('<b>Total membres</b>', styles['Normal']), str(total_membres)],
    ], colWidths=[3.0*inch, 1.0*inch])
    resume.setStyle(RLTableStyle([
        ('GRID', (0,0), (-1,-1), 0.6, colors.HexColor('#CCCCCC')),
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#F7FBFF')),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
    ]))
    story.append(resume)

    create_standard_footer(story, parametres)
    doc.build(story)
    pdf = buffer.getvalue()
    buffer.close()
    return pdf


def generate_agents_systeme_pdf():
    """Génère un PDF listant tous les agents du système avec statut de carte."""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=24, rightMargin=24, topMargin=28, bottomMargin=24)
    story = []

    parametres = get_parametres_application()
    create_standard_header(story, parametres, "AGENTS - LISTE SYSTÈME", "Administration")

    from .models import Agent
    agents = Agent.objects.order_by('nom', 'prenoms')

    styles = getSampleStyleSheet()
    story.append(Paragraph("Liste complète des agents", styles['Heading3']))

    rows = [[
        Paragraph('<b>Nom</b>', styles['Normal']),
        Paragraph('<b>Matricule</b>', styles['Normal']),
        Paragraph('<b>N° Carte</b>', styles['Normal']),
        Paragraph('<b>Carte valide</b>', styles['Normal']),
        Paragraph('<b>Caisses</b>', styles['Normal']),
    ]]
    total_agents = 0
    total_valides = 0
    for a in agents:
        etat = 'Oui' if getattr(a, 'carte_electeur_valide', False) else 'Non'
        total_agents += 1
        if etat == 'Oui':
            total_valides += 1
        rows.append([
            f"{a.nom} {a.prenoms}",
            a.matricule,
            a.numero_carte_electeur or '',
            etat,
            str(getattr(a, 'nombre_caisses', 0)),
        ])

    table = RLTable(rows, colWidths=[2.6*inch, 1.4*inch, 1.7*inch, 1.2*inch, 0.9*inch])
    table.setStyle(RLTableStyle([
        ('GRID', (0,0), (-1,-1), 0.6, colors.HexColor('#CCCCCC')),
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#F0F6FF')),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('ALIGN', (3,1), (4,-1), 'CENTER'),
    ]))
    story.append(table)
    story.append(Spacer(1, 12))

    # Résumé global
    story.append(Spacer(1, 8))
    resume = RLTable([
        [Paragraph('<b>Total cartes valides</b>', styles['Normal']), str(total_valides)],
        [Paragraph('<b>Total agents</b>', styles['Normal']), str(total_agents)],
    ], colWidths=[3.0*inch, 1.0*inch])
    resume.setStyle(RLTableStyle([
        ('GRID', (0,0), (-1,-1), 0.6, colors.HexColor('#CCCCCC')),
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#F7FBFF')),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
    ]))
    story.append(resume)

    create_standard_footer(story, parametres)
    doc.build(story)
    pdf = buffer.getvalue()
    buffer.close()
    return pdf

def generate_application_guide_pdf(buffer=None):
    """Génère un PDF 'Guide Complet de l'Application' avec les rôles et parcours clés."""
    if buffer is None:
        buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=36,
        rightMargin=36,
        topMargin=36,
        bottomMargin=40,
    )
    story = []

    # Paramètres et styles
    parametres = get_parametres_application()
    styles = getSampleStyleSheet()
    section_style = ParagraphStyle('Section', parent=styles['Heading3'], fontSize=14, spaceAfter=10, textColor=colors.HexColor('#2E86AB'))
    normal = styles['Normal']
    table_text = ParagraphStyle('TableText', parent=normal, fontSize=9, leading=12)
    table_title = ParagraphStyle('TableTitle', parent=normal, fontSize=9, leading=12)

    # En-tête
    create_standard_header(story, parametres, "GUIDE COMPLET DE L'APPLICATION", getattr(parametres, 'nom_application', '') or "")

    story.append(Paragraph("Objectif", section_style))
    story.append(Paragraph(
        "Ce document présente le fonctionnement de l'application, les rôles des utilisateurs, et les étapes clés: demandes de prêt, validations, remboursements, rapports et notifications.",
        normal,
    ))
    story.append(Spacer(1, 10))

    # Rôles et responsabilités
    story.append(Paragraph("Rôles et responsabilités", section_style))
    roles_raw = [
        ("Administrateur (Superuser)",
         "Accès complet. Valide les prêts, configure les caisses et paramètres, gère les utilisateurs, consulte les rapports et journaux d'audit."),
        ("Agent",
         "Gère les caisses assignées, enregistre les membres, accompagne les demandes et remboursements, suit les échéances."),
        ("Présidente",
         "Responsable de la caisse; valide/contrôle les opérations, suit les fonds et les prêts des membres."),
        ("Secrétaire",
         "Saisie et tenue des informations administratives de la caisse et des membres."),
        ("Trésorière",
         "Supervise les mouvements de fonds (décaissements, remboursements) et veille aux soldes."),
        ("Membre",
         "Soumet des demandes de prêt, rembourse selon les échéances, peut consulter ses informations et justificatifs (attestations)."),
    ]
    roles = [[Paragraph(f"<b>{title}</b>", table_title), Paragraph(text, table_text)] for (title, text) in roles_raw]

    role_table = RLTable(roles, colWidths=[2.4*inch, 4.4*inch], repeatRows=0)
    role_table.setStyle(RLTableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E8F4FD')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.8, colors.HexColor('#CCCCCC')),
    ]))
    story.append(role_table)
    story.append(Spacer(1, 14))

    # Parcours prêt
    story.append(Paragraph("Parcours d'un prêt", section_style))
    parcours_raw = [
        ("1. Demande", "Le membre (ou via agent) soumet une demande avec montant, durée, motif, taux d'intérêt."),
        ("2. Validation", "L'administrateur valide/rejette; en cas de validation, l'octroi est enregistré et les échéances sont générées automatiquement."),
        ("3. Remboursements", "Les paiements sont saisis; le solde se met à jour. 'Net à payer' = montant accordé + intérêts. Des reçus PDF sont générés."),
        ("4. Attestations", "A la fin, une attestation de remboursement complet peut être téléchargée."),
        ("5. Notifications", "Une cloche en haut de l'admin affiche les demandes en attente et s'actualise en temps réel."),
    ]
    parcours = [[Paragraph(f"<b>{t}</b>", table_title), Paragraph(txt, table_text)] for (t, txt) in parcours_raw]
    parcours_table = RLTable(parcours, colWidths=[1.4*inch, 5.4*inch], repeatRows=0)
    parcours_table.setStyle(RLTableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E8F5E8')),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('GRID', (0, 0), (-1, -1), 0.6, colors.HexColor('#DDDDDD')),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
    ]))
    story.append(parcours_table)
    story.append(Spacer(1, 12))

    # Navigation & fonctionnalités
    story.append(Paragraph("Navigation et fonctionnalités", section_style))
    bullets = [
        "Accueil Frontend: tableau de bord, caisses, membres, prêts, utilisateurs.",
        "Administration sécurisée: '/adminsecurelogin/' (menu latéral modernisé).",
        "PDFs: octroi de prêt, reçus de remboursement, attestations, rapports d'activité.",
        "Rapports: par type (général, financier, prêts, membres, échéances).",
        "Journaux d'audit: suivi des actions (création, modification, suppression, validation).",
        "Paramètres: identité de l'application, signatures, coordonnées et mentions légales.",
    ]
    for b in bullets:
        story.append(Paragraph(f"• {b}", normal))
    story.append(Spacer(1, 12))

    # Bonnes pratiques
    story.append(Paragraph("Bonnes pratiques & sécurité", section_style))
    for b in [
        "Utiliser des comptes personnels et des mots de passe forts.",
        "Limiter l'accès admin aux superusers et agents autorisés.",
        "Toujours télécharger et archiver les justificatifs (PDF) lors des opérations.",
        "Surveiller la cloche de notifications et les journaux d'audit.",
    ]:
        story.append(Paragraph(f"• {b}", normal))

    story.append(Spacer(1, 16))
    create_standard_footer(story, parametres)

    doc.build(story)
    pdf = buffer.getvalue()
    buffer.close()
    return pdf


def generate_prets_evaluation_pdf(date_debut=None, date_fin=None, caisse_id=None):
    """Évalue les prêts: à l'heure, en retard (avec % remboursé), non remboursés, avec synthèse.
    Filtrable par période et/ou par caisse.
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=18, rightMargin=18, topMargin=24, bottomMargin=24)
    story = []

    parametres = get_parametres_application()
    titre = "ÉVALUATION DES PRÊTS"
    sous_titre = "Toutes Caisses"
    from .models import Pret, Caisse
    caisse = None
    if caisse_id:
        try:
            caisse = Caisse.objects.filter(pk=caisse_id).first()
        except Exception:
            caisse = None
    if caisse:
        sous_titre = caisse.nom_association
    create_standard_header(story, parametres, titre, sous_titre)

    # Filtres
    qs = Pret.objects.all()
    if caisse:
        qs = qs.filter(caisse=caisse)
    if date_debut:
        qs = qs.filter(date_demande__date__gte=date_debut)
    if date_fin:
        qs = qs.filter(date_demande__date__lte=date_fin)

    # Calcul des catégories
    prets = list(qs.select_related('membre', 'caisse'))
    on_time = []
    late = []
    not_reimbursed = []

    def calc_pct(p):
        total = getattr(p, 'total_a_rembourser', 0) or 0
        restant = getattr(p, 'montant_restant', 0) or 0
        if total and total > 0:
            pct = (total - max(0, restant)) / total
        else:
            if getattr(p, 'nombre_echeances', 0):
                pct = (getattr(p, 'nombre_echeances_payees', 0) or 0) / float(p.nombre_echeances)
            else:
                pct = 0.0
        pct = 0 if pct < 0 else (1 if pct > 1 else pct)
        return round(pct * 100)

    from datetime import date as _date
    today = _date.today()
    for p in prets:
        pct = calc_pct(p)
        # Détermination délai: si date_fin_pret et statut remboursé et date_remboursement_complet <= date_fin_pret
        if p.statut == 'REMBOURSE':
            # on_time si remboursé avant/à la date d'échéance finale
            if p.date_fin_pret and p.date_remboursement_complet and p.date_remboursement_complet.date() <= p.date_fin_pret:
                on_time.append((p, pct))
            else:
                # remboursé hors délai, classer dans late comme 100%
                late.append((p, pct or 100))
        elif p.statut in ['EN_RETARD'] or (p.date_fin_pret and today > p.date_fin_pret):
            late.append((p, pct))
        elif p.statut in ['EN_ATTENTE', 'EN_ATTENTE_ADMIN', 'VALIDE', 'EN_COURS', 'BLOQUE']:
            # non remboursé (encours / bloqué)
            not_reimbursed.append((p, pct))

    styles = getSampleStyleSheet()
    section = ParagraphStyle('Sec', parent=styles['Heading3'], textColor=colors.HexColor('#2E86AB'))
    small = ParagraphStyle('Small', parent=styles['Normal'], fontSize=9)

    def add_table(title, items):
        story.append(Paragraph(title, section))
        rows = [[Paragraph('<b>Prêt</b>', small), Paragraph('<b>Membre</b>', small), Paragraph('<b>Caisse</b>', small), Paragraph('<b>% Remb.</b>', small), Paragraph('<b>Reste</b>', small)]]
        for p, pct in items:
            reste = getattr(p, 'montant_restant', 0) or 0
            rows.append([p.numero_pret, getattr(p.membre, 'nom_complet', ''), getattr(p.caisse, 'nom_association', ''), f"{pct}%", f"{reste:,.0f}".replace(',', ' ')])
        tbl = RLTable(rows, colWidths=[1.3*inch, 2.1*inch, 1.9*inch, 0.9*inch, 1.0*inch])
        tbl.setStyle(RLTableStyle([
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CCCCCC')),
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#E8F4FD')),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('ALIGN', (3,1), (4,-1), 'RIGHT'),
            ('FONTSIZE', (0,0), (-1,0), 9),
            ('FONTSIZE', (0,1), (-1,-1), 8),
        ]))
        story.append(tbl)
        story.append(Spacer(1, 10))

    # Sections
    add_table('Prêts remboursés dans le délai', on_time)
    add_table('Prêts en retard (avec % remboursé)', late)
    # Groupement non remboursés par caisse
    story.append(Paragraph('Prêts non remboursés par caisse', section))
    from collections import defaultdict
    grp = defaultdict(list)
    for p, pct in not_reimbursed:
        key = getattr(p.caisse, 'nom_association', '-')
        grp[key].append((p, pct))
    rows = [[Paragraph('<b>Caisse</b>', small), Paragraph('<b>Nombre</b>', small)]]
    for nom, lst in sorted(grp.items()):
        rows.append([nom, str(len(lst))])
    tblc = RLTable(rows, colWidths=[3.6*inch, 1.0*inch])
    tblc.setStyle(RLTableStyle([
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CCCCCC')),
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#FFF7E6')),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('ALIGN', (1,1), (1,-1), 'RIGHT'),
    ]))
    story.append(tblc)
    story.append(Spacer(1, 10))

    # Synthèse / état général
    total_prets = len(prets)
    synth_rows = [
        [Paragraph('<b>Total prêts</b>', styles['Normal']), str(total_prets)],
        [Paragraph('<b>Remboursés à l\'heure</b>', styles['Normal']), str(len(on_time))],
        [Paragraph('<b>En retard</b>', styles['Normal']), str(len(late))],
        [Paragraph('<b>Non remboursés</b>', styles['Normal']), str(len(not_reimbursed))],
    ]
    synth = RLTable(synth_rows, colWidths=[2.8*inch, 1.0*inch])
    synth.setStyle(RLTableStyle([
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CCCCCC')),
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#F0F6FF')),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
    ]))
    story.append(Paragraph('État général', section))
    story.append(synth)

    create_standard_footer(story, parametres)
    doc.build(story)
    pdf = buffer.getvalue()
    buffer.close()
    return pdf


def generate_prets_par_motif_pdf(motif=None, date_debut=None, date_fin=None):
    """PDF global (admin) listant les prêts filtrés par motif avec % remboursé."""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=18, rightMargin=18, topMargin=24, bottomMargin=24)
    story = []

    parametres = get_parametres_application()
    titre = "PRÊTS PAR MOTIF"
    sous_titre = motif.capitalize() if motif else 'Tous motifs'
    create_standard_header(story, parametres, titre, sous_titre)

    from .models import Pret
    from django.db.models import Q
    qs = Pret.objects.select_related('membre','caisse')
    if motif:
        qs = qs.filter(Q(motif__iexact=motif) | Q(motif__icontains=motif))
    if date_debut:
        qs = qs.filter(date_demande__date__gte=date_debut)
    if date_fin:
        qs = qs.filter(date_demande__date__lte=date_fin)

    styles = getSampleStyleSheet()
    small = ParagraphStyle('Small', parent=styles['Normal'], fontSize=9)
    rows = [[Paragraph('<b>N° Prêt</b>', small), Paragraph('<b>Membre</b>', small), Paragraph('<b>Caisse</b>', small), Paragraph('<b>% Remb.</b>', small)]]

    def pct_pret(p):
        total = getattr(p, 'total_a_rembourser', 0) or 0
        restant = getattr(p, 'montant_restant', 0) or 0
        if total > 0:
            v = (total - max(0, restant)) / total
        else:
            v = 0
        v = 0 if v < 0 else (1 if v > 1 else v)
        return round(v * 100)

    for p in qs:
        rows.append([p.numero_pret, getattr(p.membre, 'nom_complet', ''), getattr(p.caisse, 'nom_association', ''), f"{pct_pret(p)}%"])

    tbl = RLTable(rows, colWidths=[1.3*inch, 2.2*inch, 2.0*inch, 0.9*inch])
    tbl.setStyle(RLTableStyle([
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CCCCCC')),
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#E8F4FD')),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('ALIGN', (3,1), (3,-1), 'RIGHT'),
        ('FONTSIZE', (0,0), (-1,0), 9),
        ('FONTSIZE', (0,1), (-1,-1), 8),
    ]))
    story.append(tbl)

    create_standard_footer(story, parametres)
    doc.build(story)
    pdf = buffer.getvalue()
    buffer.close()
    return pdf
def export_rapport_excel(rapport):
    """Exporte un rapport en format Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.chart import BarChart, Reference
        from openpyxl.utils.dataframe import dataframe_to_rows
        import pandas as pd
        from io import BytesIO
        from django.http import HttpResponse
    except ImportError:
        # Fallback si openpyxl n'est pas installé
        return export_rapport_csv(rapport)
    
    # Créer un nouveau classeur Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Rapport_{rapport.type_rapport}"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
    title_font = Font(bold=True, size=16, color="2E86AB")
    
    # En-tête
    ws['A1'] = f"RAPPORT {rapport.type_rapport.upper()}"
    ws['A1'].font = title_font
    ws.merge_cells('A1:F1')
    
    if rapport.caisse:
        ws['A2'] = f"Caisse: {rapport.caisse.nom_association}"
    else:
        ws['A2'] = "Toutes les Caisses"
    
    if rapport.date_debut and rapport.date_fin:
        ws['A3'] = f"Période: {rapport.date_debut.strftime('%d/%m/%Y')} → {rapport.date_fin.strftime('%d/%m/%Y')}"
    
    ws['A5'] = f"Généré le: {rapport.date_generation.strftime('%d/%m/%Y %H:%M') if rapport.date_generation else 'Non généré'}"
    ws['A6'] = f"Par: {rapport.genere_par.get_full_name() if rapport.genere_par else 'Non défini'}"
    
    # Données du rapport
    if rapport.donnees:
        data = rapport.donnees
        row = 8
        
        # Parcourir les données et les organiser en tableaux
        for section_name, section_data in data.items():
            if isinstance(section_data, dict):
                # En-tête de section
                ws[f'A{row}'] = section_name.replace('_', ' ').title()
                ws[f'A{row}'].font = Font(bold=True, size=14, color="2E86AB")
                row += 1
                
                # Tableau des données
                if section_data:
                    # En-têtes du tableau
                    headers = list(section_data.keys())
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=row, column=col, value=header.replace('_', ' ').title())
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal="center")
                    
                    row += 1
                    
                    # Données
                    for col, value in enumerate(section_data.values(), 1):
                        cell = ws.cell(row=row, column=col, value=str(value))
                        cell.alignment = Alignment(horizontal="left")
                    
                    row += 2
                else:
                    ws[f'A{row}'] = "Aucune donnée"
                    row += 2
                    
            elif isinstance(section_data, list) and section_data:
                # Liste de données
                ws[f'A{row}'] = section_name.replace('_', ' ').title()
                ws[f'A{row}'].font = Font(bold=True, size=14, color="2E86AB")
                row += 1
                
                if section_data and isinstance(section_data[0], dict):
                    # En-têtes du tableau
                    headers = list(section_data[0].keys())
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=row, column=col, value=header.replace('_', ' ').title())
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal="center")
                    
                    row += 1
                    
                    # Données
                    for item in section_data:
                        for col, value in enumerate(item.values(), 1):
                            cell = ws.cell(row=row, column=col, value=str(value))
                            cell.alignment = Alignment(horizontal="left")
                        row += 1
                    
                    row += 2
                else:
                    # Liste simple
                    for item in section_data:
                        ws[f'A{row}'] = str(item)
                        row += 1
                    row += 2
    
    # Ajuster la largeur des colonnes
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Créer la réponse HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="rapport_{rapport.type_rapport}_{rapport.pk}.xlsx"'
    
    # Sauvegarder dans un buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response.write(buffer.getvalue())
    
    return response

def export_rapports_excel(queryset):
    """Exporte plusieurs rapports en format Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        from django.http import HttpResponse
    except ImportError:
        # Fallback si openpyxl n'est pas installé
        return export_rapports_csv(queryset)
    
    # Créer un nouveau classeur Excel
    wb = openpyxl.Workbook()
    
    # Supprimer la feuille par défaut
    wb.remove(wb.active)
    
    # Créer une feuille par rapport
    for rapport in queryset:
        if rapport.donnees:
            ws = wb.create_sheet(f"Rapport_{rapport.type_rapport}_{rapport.pk}")
            
            # En-tête
            ws['A1'] = f"RAPPORT {rapport.type_rapport.upper()}"
            ws['A1'].font = Font(bold=True, size=16, color="2E86AB")
            ws.merge_cells('A1:F1')
            
            if rapport.caisse:
                ws['A2'] = f"Caisse: {rapport.caisse.nom_association}"
            else:
                ws['A2'] = "Toutes les Caisses"
            
            if rapport.date_debut and rapport.date_fin:
                ws['A3'] = f"Période: {rapport.date_debut.strftime('%d/%m/%Y')} → {rapport.date_fin.strftime('%d/%m/%Y')}"
            
            ws['A5'] = f"Généré le: {rapport.date_generation.strftime('%d/%m/%Y %H:%M') if rapport.date_generation else 'Non généré'}"
            
            # Données simplifiées
            row = 7
            for key, value in rapport.donnees.items():
                if isinstance(value, (dict, list)):
                    ws[f'A{row}'] = f"{key}: {len(value)} éléments"
                else:
                    ws[f'A{row}'] = f"{key}: {value}"
                row += 1
    
    # Créer la réponse HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="rapports_multiples_{queryset.count()}.xlsx"'
    
    # Sauvegarder dans un buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response.write(buffer.getvalue())
    
    return response

def export_rapport_csv(rapport):
    """Exporte un rapport en format CSV"""
    import csv
    from io import StringIO
    from django.http import HttpResponse
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="rapport_{rapport.type_rapport}_{rapport.pk}.csv"'
    
    # Encodage UTF-8 avec BOM pour Excel
    response.write('\ufeff')
    
    writer = csv.writer(response)
    
    # En-tête
    writer.writerow([f"RAPPORT {rapport.type_rapport.upper()}"])
    if rapport.caisse:
        writer.writerow([f"Caisse: {rapport.caisse.nom_association}"])
    else:
        writer.writerow(["Toutes les Caisses"])
    
    if rapport.date_debut and rapport.date_fin:
        writer.writerow([f"Période: {rapport.date_debut.strftime('%d/%m/%Y')} → {rapport.date_fin.strftime('%d/%m/%Y')}"])
    
    writer.writerow([f"Généré le: {rapport.date_generation.strftime('%d/%m/%Y %H:%M') if rapport.date_generation else 'Non généré'}"])
    writer.writerow([f"Par: {rapport.genere_par.get_full_name() if rapport.genere_par else 'Non défini'}"])
    writer.writerow([])
    
    # Données du rapport
    if rapport.donnees:
        data = rapport.donnees
        
        for section_name, section_data in data.items():
            writer.writerow([section_name.replace('_', ' ').title()])
            
            if isinstance(section_data, dict):
                # En-têtes
                headers = list(section_data.keys())
                writer.writerow([h.replace('_', ' ').title() for h in headers])
                
                # Données
                values = list(section_data.values())
                writer.writerow([str(v) for v in values])
                
            elif isinstance(section_data, list) and section_data:
                if isinstance(section_data[0], dict):
                    # En-têtes
                    headers = list(section_data[0].keys())
                    writer.writerow([h.replace('_', ' ').title() for h in headers])
                    
                    # Données
                    for item in section_data:
                        writer.writerow([str(item.get(h, '')) for h in headers])
                else:
                    # Liste simple
                    for item in section_data:
                        writer.writerow([str(item)])
            
            writer.writerow([])
    
    return response

def export_rapports_csv(queryset):
    """Exporte plusieurs rapports en format CSV"""
    import csv
    from io import StringIO
    from django.http import HttpResponse
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="rapports_multiples_{queryset.count()}.csv"'
    
    # Encodage UTF-8 avec BOM pour Excel
    response.write('\ufeff')
    
    writer = csv.writer(response)
    
    # En-tête général
    writer.writerow(["EXPORT MULTIPLE DE RAPPORTS"])
    writer.writerow([f"Nombre de rapports: {queryset.count()}"])
    writer.writerow([])
    
    # Données par rapport
    for rapport in queryset:
        writer.writerow([f"RAPPORT {rapport.type_rapport.upper()}"])
        writer.writerow([f"ID: {rapport.pk}"])
        
        if rapport.caisse:
            writer.writerow([f"Caisse: {rapport.caisse.nom_association}"])
        else:
            writer.writerow(["Caisse: Toutes les Caisses"])
        
        if rapport.date_debut and rapport.date_fin:
            writer.writerow([f"Période: {rapport.date_debut.strftime('%d/%m/%Y')} → {rapport.date_fin.strftime('%d/%m/%Y')}"])
        
        writer.writerow([f"Statut: {rapport.statut}"])
        writer.writerow([f"Généré le: {rapport.date_generation.strftime('%d/%m/%Y %H:%M') if rapport.date_generation else 'Non généré'}"])
        
        # Résumé des données
        if rapport.donnees:
            writer.writerow(["Résumé des données:"])
            for key, value in rapport.donnees.items():
                if isinstance(value, (dict, list)):
                    writer.writerow([f"  {key}: {len(value)} éléments"])
                else:
                    writer.writerow([f"  {key}: {value}"])
        
        writer.writerow([])
        writer.writerow(["---"])
        writer.writerow([])
    
    return response

def generate_fiche_paie_pdf(fiche_paie):
    """Génère le PDF de la fiche de paie d'un agent"""
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.platypus import PageBreak
    from io import BytesIO
    from django.core.files.base import ContentFile
    import os
    
    # Créer le buffer pour le PDF
    buffer = BytesIO()
    
    # Créer le document PDF
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    story = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1,  # Centré
        textColor=colors.darkblue
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Heading2'],
        fontSize=14,
        spaceAfter=20,
        alignment=1,  # Centré
        textColor=colors.darkblue
    )
    
    header_style = ParagraphStyle(
        'Header',
        parent=styles['Heading3'],
        fontSize=12,
        spaceAfter=10,
        textColor=colors.darkblue
    )
    
    # En-tête de la fiche de paie
    story.append(Paragraph("FICHE DE PAIE", title_style))
    story.append(Paragraph(f"Période: {fiche_paie.periode}", subtitle_style))
    story.append(Spacer(1, 20))
    
    # Informations de l'agent
    story.append(Paragraph("INFORMATIONS DE L'AGENT", header_style))
    
    agent_info = [
        ["Nom et Prénoms:", fiche_paie.nom_agent],
        ["Matricule:", fiche_paie.matricule],
        ["Poste:", fiche_paie.poste],
        ["Période:", fiche_paie.periode]
    ]
    
    agent_table = Table(agent_info, colWidths=[2*inch, 4*inch])
    agent_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(agent_table)
    story.append(Spacer(1, 20))
    
    # Détails du salaire
    story.append(Paragraph("DÉTAILS DU SALAIRE", header_style))
    
    # Récapitulatif des gains
    gains_info = [
        ["Salaire de base:", f"{fiche_paie.salaire_base:,.0f} FCFA"],
        ["Bonus caisses:", f"{fiche_paie.bonus_caisses:,.0f} FCFA"],
        ["Prime de performance:", f"{fiche_paie.prime_performance:,.0f} FCFA"],
        ["", ""],
        ["TOTAL BRUT:", f"{fiche_paie.total_brut:,.0f} FCFA"]
    ]
    
    gains_table = Table(gains_info, colWidths=[3*inch, 3*inch])
    gains_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.black),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 4), (1, 4), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 4), (1, 4), 12)
    ]))
    
    story.append(gains_table)
    story.append(Spacer(1, 20))
    
    # Déductions et total net
    story.append(Paragraph("DÉDUCTIONS ET TOTAL NET", header_style))
    
    net_info = [
        ["Déductions:", f"{fiche_paie.deductions:,.0f} FCFA"],
        ["", ""],
        ["TOTAL NET À PAYER:", f"{fiche_paie.total_net:,.0f} FCFA"]
    ]
    
    net_table = Table(net_info, colWidths=[3*inch, 3*inch])
    net_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.black),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 2), (1, 2), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 2), (1, 2), 12),
        ('BACKGROUND', (0, 2), (1, 2), colors.lightblue)
    ]))
    
    story.append(net_table)
    story.append(Spacer(1, 20))
    
    # Informations sur les caisses
    if fiche_paie.nombre_nouvelles_caisses > 0:
        story.append(Paragraph("INFORMATIONS SUR LES CAISSES", header_style))
        
        caisses_info = [
            ["Nombre de nouvelles caisses créées:", str(fiche_paie.nombre_nouvelles_caisses)],
            ["Bonus par caisse:", "5 000 FCFA"],
            ["Total bonus caisses:", f"{fiche_paie.bonus_caisses:,.0f} FCFA"]
        ]
        
        caisses_table = Table(caisses_info, colWidths=[3*inch, 3*inch])
        caisses_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.black),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(caisses_table)
        story.append(Spacer(1, 20))
    
    # Signature
    story.append(Paragraph("SIGNATURES", header_style))
    
    signature_info = [
        ["Agent:", "_________________", "Date:", "_________________"],
        ["", "", "", ""],
        ["Responsable RH:", "_________________", "Date:", "_________________"],
        ["", "", "", ""],
        ["Directeur Général:", "_________________", "Date:", "_________________"]
    ]
    
    signature_table = Table(signature_info, colWidths=[1.5*inch, 2*inch, 1.5*inch, 2*inch])
    signature_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(signature_table)
    
    # Générer le PDF
    doc.build(story)
    
    # Récupérer le contenu du buffer
    pdf_content = buffer.getvalue()
    buffer.close()
    
    # Créer le nom du fichier
    filename = f"fiche_paie_{fiche_paie.matricule}_{fiche_paie.mois}_{fiche_paie.annee}.pdf"
    
    # Retourner le fichier
    return ContentFile(pdf_content, name=filename)
